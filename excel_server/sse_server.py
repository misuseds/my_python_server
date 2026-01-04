# e:\code\my_python_server\sse\sse_server.py
import asyncio
from flask import Flask, request, jsonify
from openpyxl import load_workbook
import os

app = Flask(__name__)

def read_excel_data(filepath, sheet_name=None):
    """直接读取 Excel 数据，不使用 MCP"""
    try:
        # 检查文件是否存在
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")
        
        # 加载工作簿
        workbook = load_workbook(filepath, data_only=True)
        
        # 处理 sheet_name 参数
        if sheet_name is None:
            worksheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            worksheet = workbook[sheet_name]
        
        # 读取所有数据
        data = [list(row) for row in worksheet.iter_rows(values_only=True)]
        
        # 获取列标题
        headers = [cell.value for cell in worksheet[1]]
        
        return {
            "headers": headers,
            "data": data,
            "sheet_name": worksheet.title,
            "total_rows": worksheet.max_row,
            "total_columns": worksheet.max_column
        }
    except Exception as e:
        raise Exception(f"读取Excel文件时出错: {str(e)}")


@app.route('/read_excel', methods=['GET'])
def read_excel_endpoint():
    """Flask 路由，接受 filepath 和可选的 sheet_name 参数并返回读取结果"""
    filepath = request.args.get('filepath')
    sheet_name = request.args.get('sheet_name')
    
    if not filepath:
        return jsonify({"error": "缺少 filepath 参数"}), 400
    
    try:
        result = read_excel_data(filepath, sheet_name)
        return jsonify({"result": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/routes')
def show_routes():
    routes = []
    for rule in app.url_map.iter_rules():
        # 获取端点函数的docstring
        endpoint_func = app.view_functions.get(rule.endpoint)
        docstring = endpoint_func.__doc__.strip() if endpoint_func and endpoint_func.__doc__ else "无描述"
        
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule),
            'description': docstring
        })
    return {'routes': routes}

import argparse

# 解析命令行参数
parser = argparse.ArgumentParser()
parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
args = parser.parse_args()

# 使用指定的端口运行 Flask 应用
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=args.port, debug=True)