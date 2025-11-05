import json
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import sys
import tkinter as tk
from tkinter import filedialog
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from llm_server.llm_server import LLMService

llm_service = LLMService()

class ExcelProcessor:
    def __init__(self):
        pass
    
    def select_excel_files(self):
        """
        弹出文件选择窗口选择Excel文件
        """
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_types = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        file_paths = filedialog.askopenfilenames(
            title="请选择一个或多个Excel文件",
            filetypes=file_types
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return list(file_paths)
    
    def read_excel_content(self, file_path):
        """
        读取Excel文件内容并转换为文本形式
        
        Args:
            file_path (str): Excel文件路径
            
        Returns:
            str: Excel内容的文本表示
        """
        try:
            # 使用pandas读取Excel文件的所有工作表
            excel_data = pd.read_excel(file_path, sheet_name=None)
            
            content_text = ""
            for sheet_name, df in excel_data.items():
                content_text += f"工作表: {sheet_name}\n"
                content_text += df.to_string(index=False)
                content_text += "\n\n"
                
            return content_text
        except Exception as e:
            logger.error(f"读取Excel文件 {file_path} 时出错: {str(e)}")
            return ""
    
    def parse_vlm_response(self, response_data):
        """
        解析 LLM 返回的结构化数据
        """
        try:
            result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            if result.startswith("```json"):
                result = result[7:]
            if result.startswith("```"):
                result = result[3:]
            if result.endswith("```"):
                result = result[:-3]
            
            result = result.strip()
            parsed = json.loads(result)
            
            if "processed_data" in parsed:
                return {"processed_data": parsed["processed_data"]}
            else:
                return {}
        except Exception as e:
            print(f"解析失败: {e}")
            return {}
    
    def save_processed_data_to_excel(self, data, original_file_path):
        """
        将处理后的数据保存为新的Excel文件
        
        Args:
            data (dict): 处理后的结构化数据
            original_file_path (str): 原始文件路径
        """
        # 生成新文件名，在原文件名后加上"_processed"
        file_dir = os.path.dirname(original_file_path)
        file_name = os.path.splitext(os.path.basename(original_file_path))[0]
        file_ext = os.path.splitext(original_file_path)[1]
        new_file_path = os.path.join(file_dir, f"{file_name}_processed{file_ext}")
        
        # 防止文件名重复
        base_path, ext = os.path.splitext(new_file_path)
        counter = 1
        final_output_path = new_file_path
        while os.path.exists(final_output_path):
            final_output_path = f"{base_path}_{counter}{ext}"
            counter += 1
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Data"
        
        # 设置表头
        headers = ["规格", "长度", "数量"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 填充数据
        if "processed_data" in data and data["processed_data"]:
            for row_idx, item in enumerate(data["processed_data"], 2):
                ws.cell(row=row_idx, column=1, value=item.get("规格", ""))
                ws.cell(row=row_idx, column=2, value=item.get("长度", ""))
                ws.cell(row=row_idx, column=3, value=item.get("数量", ""))
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存文件
        wb.save(final_output_path)
        print(f"处理后的Excel文件已保存至: {final_output_path}")
        return final_output_path
    
    def process_excel_file(self, file_path):
        """
        主函数：读取Excel文件，调用LLM处理，并生成新的Excel文件
        
        Args:
            file_path (str): 输入Excel文件路径
        """
        try:
            # 读取Excel内容
            excel_content = self.read_excel_content(file_path)
            if not excel_content:
                raise Exception("无法读取Excel文件内容")
            
            # 构建提示消息
            messages = [
                {
                    "role": "user",
                    "content": f'''请分析以下Excel表格内容，并提取出包含"规格"、"长度"、"数量"信息的数据项。
请将提取的信息按照以下JSON格式返回：

输出格式要求：
{{
    "processed_data": [
        {{
            "规格": "示例规格值",
            "长度": "示例长度值",
            "数量": "示例数量值"
        }}
    ]
}}

Excel内容：
{excel_content}

注意事项：
1. 只提取包含完整"规格"、"长度"、"数量"信息的条目
2. 保持原有单位（如有）
3. 如有计算公式需求，请在相应字段中体现
4. 确保返回有效的JSON格式'''
                }
            ]
            
            # 调用LLM服务
            result = llm_service.create(messages)
            print("LLM 返回:", result)
            processed_data = self.parse_vlm_response(result)
            
            # 保存处理后的数据到新的Excel文件
            self.save_processed_data_to_excel(processed_data, file_path)
            
        except Exception as e:
            print(f"处理文件 '{file_path}' 时出现错误: {e}")
    
    def batch_process_excel_files(self, file_paths):
        """
        批量处理多个Excel文件
        
        Args:
            file_paths (list): Excel文件路径列表
        """
        if not file_paths:
            print("未选择任何文件")
            return
        
        for i, file_path in enumerate(file_paths, 1):
            print(f"\n正在处理第 {i}/{len(file_paths)} 个文件: {file_path}")
            
            if not os.path.exists(file_path):
                print(f"警告: 文件 '{file_path}' 不存在，跳过处理")
                continue
                
            try:
                self.process_excel_file(file_path)
                print(f"完成处理: {file_path}")
            except Exception as e:
                print(f"处理文件 '{file_path}' 时出现错误: {e}")

def main():
    # 创建处理器实例
    processor = ExcelProcessor()
    
    # 选择Excel文件进行批量处理
    file_paths = processor.select_excel_files()
    
    if not file_paths:
        print("未选择文件，程序退出")
        sys.exit(0)
        
    try:
        processor.batch_process_excel_files(file_paths)
        print("\n所有Excel文件处理完成！")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")

if __name__ == '__main__':
    main()