# txt_to_excel.py
import json
import os
import sys
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import logging
import tkinter as tk
from tkinter import filedialog

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 添加项目路径到系统路径
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from llm_server.llm_class import LLMService

llm_service = LLMService()

class TextToExcelConverter:
    def __init__(self):
        # 定义提取字段配置
        self.extraction_fields = [
            {"key": "drawing_number", "label": "P7764", "display_name": "图号"},
            {"key": "specification", "label": "PL厚度*宽度*长度（类似PL8*504*1000）", "display_name": "规格"},
            {"key": "quantity", "label": "5", "display_name": "数量"},
            {"key": "weight", "label": "1000", "display_name": "重量"},
            {"key": "material", "label": "不锈钢", "display_name": "材料"},
            {"key": "remark", "label": "不锈钢板", "display_name": "备注"}
        ]
    
    def select_txt_file(self):
        """
        弹出文件选择窗口选择txt文件
        """
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_types = [
            ("Text files", "*.txt"),
            ("All files", "*.*")
        ]
        
        file_path = filedialog.askopenfilename(
            title="请选择一个TXT文件",
            filetypes=file_types
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return file_path
    
    def read_text_file(self, file_path):
        """
        读取文本文件内容
        
        Args:
            file_path (str): 文本文件路径
            
        Returns:
            str: 文件内容
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"成功读取文件: {file_path}")
            return content
        except Exception as e:
            logger.error(f"读取文件失败: {str(e)}")
            raise
    
    def parse_llm_response(self, response_data):
        """
        解析 LLM 返回的结构化数据
        """
        try:
            result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            if result.startswith("```json"):
                result = result[7:]
            if result.startswith("```"):
                result = result[3:]
            if result.endswith("```"):
                result = result[:-3]
            
            result = result.strip()
            parsed = json.loads(result)
            
            if "drawing_info" in parsed:
                return {"drawing_info": parsed["drawing_info"]}
            else:
                return {}
        except Exception as e:
            print(f"解析失败: {e}")
            return {}
    
    def _convert_specification_format(self, value):
        """
        转换规格格式，添加PL前缀（如果没有）并将x替换为*，去除空格，保留一位小数
        """
        import re
        
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 去除所有空格
            value = value.replace(' ', '')
            
            # 替换所有的"x"为"*"
            value = value.replace('x', '*').replace('X', '*')
            
            # 处理数字，保留一位小数
            # 使用正则表达式找到所有数字部分并格式化
            def format_number(match):
                try:
                    num = float(match.group())
                    # 如果是整数，不保留小数点
                    if num.is_integer():
                        return str(int(num))
                    else:
                        # 保留一位小数
                        return f"{num:.1f}"
                except:
                    return match.group()
            
            # 匹配数字（包括小数）
            value = re.sub(r'\d+\.?\d*', format_number, value)
            
            # 如果不以PL开头，则添加PL前缀
            if not value.upper().startswith('PL'):
                value = 'PL' + value
                
            return value
        
        return value

    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 处理Unicode数字字符（如①②③等）
            try:
                unicode_digits = {
                    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
                    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10
                }
                if value in unicode_digits:
                    return unicode_digits[value]
            except:
                pass
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value

    def save_data_to_excel(self, data, output_path="output.xlsx", text_filename=None):
        """
        将结构化数据保存为Excel文件
        
        Args:
            data (dict): 结构化数据
            output_path (str): 输出Excel文件路径
            text_filename (str): 原始文本文件名，用于填入Excel前三行合并单元格
        """
        # 防止文件名重复
        base_path, ext = os.path.splitext(output_path)
        counter = 1
        final_output_path = output_path
        while os.path.exists(final_output_path):
            final_output_path = f"{base_path}_{counter}{ext}"
            counter += 1
        
        wb = Workbook()
        ws = wb.active
        
        # 将原来的三行标题改为一行
        if text_filename:
            ws["A1"] = text_filename
            ws.merge_cells("A1:E1")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(size=14, bold=True)
            ws.row_dimensions[1].height = 25
            start_row = 2
        else:
            start_row = 1
        
        rows = []
        row_count = 0  # 记录数据行数
        
        if "drawing_info" in data and data["drawing_info"]:
            rows.append(["图纸信息", "", "", "", ""])  # 标题行
            row_count += 1
            
            # 动态提取所有字段名作为表头
            all_fields = set()
            for item in data["drawing_info"]:
                all_fields.update(item.keys())
            
            # 创建显示名称映射
            display_names = {field["key"]: field["display_name"] for field in self.extraction_fields}
            
            # 创建表头行（中文）
            header_row = []
            field_order = []  # 保持列顺序
            
            # 按预定义顺序添加字段
            for field_config in self.extraction_fields:
                field_key = field_config["key"]
                if field_key in all_fields:
                    header_row.append(display_names.get(field_key, field_key))
                    field_order.append(field_key)
            
            # 添加其他未预定义的字段
            for field in all_fields:
                if field not in field_order:
                    header_row.append(field)  # 对于未预定义字段，直接使用字段名
                    field_order.append(field)
                    
            rows.append(header_row)
            row_count += 1
            
            # 填充数据行
            for item in data["drawing_info"]:
                data_row = []
                for field in field_order:
                    cell_value = item.get(field, "")
                    # 特殊处理 specification 字段
                    if field == "specification":
                        cell_value = self._convert_specification_format(cell_value)
                    # 尝试将值转换为数字
                    final_value = self._convert_to_number_if_possible(cell_value)
                    data_row.append(final_value)
                rows.append(data_row)
                row_count += 1
        
        # 写入Excel
        for i, row in enumerate(rows, start=start_row):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
            # 设置每行行高为25
            ws.row_dimensions[i].height = 25
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大宽度
        
        wb.save(final_output_path)
        print(f"Excel 文件已保存至: {final_output_path}")
        return final_output_path   

    def analyze_text_and_generate_excel(self, text_path):
        """
        主函数：调用 LLM 分析文本，并生成 Excel 文件
        
        Args:
            text_path (str): 输入文本文件路径
        """
        try:
            # 读取文本内容
            text_content = self.read_text_file(text_path)
            
            # 获取文件目录和文件名（不含扩展名）
            file_dir = os.path.dirname(text_path)
            file_name = os.path.splitext(os.path.basename(text_path))[0]
            output_path = os.path.join(file_dir, f"{file_name}.xlsx")
            
            # 根据当前字段配置动态生成提示语句
            labels = [field["display_name"] for field in self.extraction_fields]
            labels_str = "、".join(labels)
            instruction = f"请从以下文本中提取信息，尽可能提取出{labels_str}等信息"
            
            # 动态构建JSON示例中的字段部分
            field_lines = []
            for field in self.extraction_fields:
                field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
            
            json_fields_block = ",\n".join(field_lines)
            
            messages = [
                {
                    "role": "user",
                    "content": f'''文件名: {os.path.basename(text_path)}
{instruction}，并以 JSON 格式返回。
输出格式要求：
{{
    "drawing_info": [
        {{
{json_fields_block}
        }}
    ]
}}

以下是需要分析的文本内容：
{text_content}'''
                }
            ]
            
            result = llm_service.create(messages)
            print("LLM 返回:", result)
            extracted_data = self.parse_llm_response(result)
            self.save_data_to_excel(extracted_data, output_path, os.path.basename(text_path))
            
        except Exception as e:
            print("错误:", str(e))
            raise

def main():
    converter = TextToExcelConverter()
    
    # 弹窗选择文件
    text_file_path = converter.select_txt_file()
    
    if not text_file_path:
        print("未选择文件，程序退出")
        sys.exit(0)
    
    if not os.path.exists(text_file_path):
        print(f"错误: 文件 '{text_file_path}' 不存在")
        sys.exit(1)
        
    try:
        converter.analyze_text_and_generate_excel(text_file_path)
        print("\n文本处理完成，Excel文件已生成！")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()