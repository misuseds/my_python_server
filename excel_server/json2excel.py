# file: json_to_excel_service.py
import pandas as pd
import json
from flask import Flask, request, jsonify, send_file
import os
import logging
from openpyxl import load_workbook

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)


def save_json_to_excel(json_data, filename="output.xlsx", dxf_filename=None):
    """
    将JSON数据保存为Excel文件
    
    Args:
        json_data (list/dict): JSON格式的数据
        filename (str): 输出的Excel文件名
        dxf_filename (str): DXF文件名，用于填入Excel前三行合并单元格
    
    Returns:
        str: 生成的Excel文件路径
    """
    try:
        # 如果输入是JSON字符串，则解析为Python对象
        if isinstance(json_data, str):
            json_data = json.loads(json_data)
        
        # 如果是字典且包含列表数据，需要特殊处理
        if isinstance(json_data, dict):
            # 如果字典中的值是列表，使用第一个列表
            if any(isinstance(v, list) for v in json_data.values()):
                # 找到第一个列表值作为数据源
                for value in json_data.values():
                    if isinstance(value, list):
                        df = pd.DataFrame(value)
                        break
            else:
                # 如果没有列表，将整个字典转换为单行DataFrame
                df = pd.DataFrame([json_data])
        elif isinstance(json_data, list):
            # 直接用列表创建DataFrame
            df = pd.DataFrame(json_data)
        else:
            raise ValueError("不支持的数据格式")
            
        # 保存为Excel文件
        output_path = os.path.abspath(filename)
        df.to_excel(output_path, index=False)
        
        # 如果提供了文件名，则在前三行合并单元格并填入文件名
        if dxf_filename:
            workbook = load_workbook(output_path)
            worksheet = workbook.active
            
            # 获取最大列数
            max_column = worksheet.max_column
            if max_column > 0:
                # 在第一行之前插入3行
                worksheet.insert_rows(1, 3)
                
                # 合并新插入的前三行的单元格
                merge_range = f"A1:{chr(64 + max_column) if max_column <= 26 else 'Z' + chr(64 + max_column - 26)}3" if max_column > 26 else f"A1:{chr(64 + max_column)}3"
                if max_column <= 26:
                    merge_range = f"A1:{chr(64 + max_column)}3"
                else:
                    # 处理超过26列的情况
                    first_char = chr(64 + (max_column - 1) // 26) if (max_column - 1) // 26 > 0 else ""
                    second_char = chr(64 + ((max_column - 1) % 26) + 1)
                    last_column = first_char + second_char
                    merge_range = f"A1:{last_column}3"
                    
                worksheet.merge_cells(merge_range)
                
                # 在合并的单元格中填入文件名
                worksheet["A1"] = dxf_filename
                
                # 设置居中对齐
                from openpyxl.styles import Alignment
                alignment = Alignment(horizontal="center", vertical="center")
                worksheet["A1"].alignment = alignment
                
                # 可选：设置字体大小
                from openpyxl.styles import Font
                font = Font(size=14, bold=True)
                worksheet["A1"].font = font
            
            # 保存修改后的Excel文件
            workbook.save(output_path)
        
        logger.info(f"Excel文件已保存至: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"保存Excel文件时出错: {str(e)}")
        raise e



@app.route('/json_to_excel', methods=['POST'])
def convert_json_to_excel():
    """
    接收JSON数据并转换为Excel文件的API接口
    
    ---
    parameters:
      - name: json_data
        in: body
        required: true
        description: JSON格式的数据
      - name: filename
        in: query
        type: string
        required: false
        default: output.xlsx
        description: 输出的Excel文件名
      - name: dxf_filename
        in: query
        type: string
        required: false
        description: DXF文件名，将填入Excel前三行合并单元格
    responses:
      200:
        description: 成功生成Excel文件
        content:
          application/vnd.openxmlformats-officedocument.spreadsheetml.sheet:
            schema:
              type: string
              format: binary
      400:
        description: 请求参数错误
      500:
        description: 服务器内部错误
    """
    try:
        # 获取请求体中的JSON数据
        if request.is_json:
            json_data = request.get_json()
        else:
            return jsonify({"error": "请求必须包含JSON数据"}), 400
            
        # 获取查询参数中的文件名和DXF文件名
        filename = request.args.get('filename', 'output.xlsx')
        dxf_filename = request.args.get('dxf_filename')
        
        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        # 调用转换函数
        file_path = save_json_to_excel(json_data, filename, dxf_filename)
        
        # 返回文件下载
        return send_file(file_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"处理请求时出错: {str(e)}")
        return jsonify({"error": f"处理过程中出错: {str(e)}"}), 500

@app.route('/save_structured_data', methods=['POST'])
def save_structured_data():
    """
    专门用于保存getthickpart服务中LLM返回的结构化数据为Excel文件
    
    ---
    parameters:
      - name: data
        in: body
        required: true
        description: LLM返回的结构化数据
      - name: filename
        in: query
        type: string
        required: false
        default: structured_data.xlsx
        description: 输出的Excel文件名
      - name: dxf_filename
        in: query
        type: string
        required: false
        description: DXF文件名，将填入Excel前三行合并单元格
    responses:
      200:
        description: 成功保存Excel文件
      400:
        description: 请求参数错误
      500:
        description: 服务器内部错误
    """
    try:
        # 获取请求体中的JSON数据
        if request.is_json:
            request_data = request.get_json()
        else:
            return jsonify({"error": "请求必须包含JSON数据"}), 400
            
        # 提取结构化数据
        if 'structured_data' in request_data:
            structured_data = request_data['structured_data']
        else:
            structured_data = request_data
            
        # 获取查询参数中的文件名和DXF文件名
        filename = request.args.get('filename')
        dxf_filename = request.args.get('dxf_filename')
        if not filename:
            # 如果没有指定文件名，使用默认命名规则
            filename = 'structured_data.xlsx'
        
        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        # 调用转换函数
        file_path = save_json_to_excel(structured_data, filename, dxf_filename)
        
        # 返回成功信息
        return jsonify({
            "status": "success",
            "message": f"数据已成功保存为Excel文件: {filename}",
            "file_path": file_path
        })
        
    except Exception as e:
        logger.error(f"处理请求时出错: {str(e)}")
        return jsonify({"error": f"处理过程中出错: {str(e)}"}), 500

@app.route('/json_file_to_excel', methods=['GET'])
def convert_json_file_to_excel():
    """
    读取本地JSON文件并转换为Excel文件的API接口
    
    ---
    parameters:
      - name: filepath
        in: query
        type: string
        required: true
        description: 本地JSON文件的路径
      - name: filename
        in: query
        type: string
        required: false
        default: output.xlsx
        description: 输出的Excel文件名
      - name: dxf_filename
        in: query
        type: string
        required: false
        description: DXF文件名，将填入Excel前三行合并单元格
    responses:
      200:
        description: 成功生成Excel文件
        content:
          application/vnd.openxmlformats-officedocument.spreadsheetml.sheet:
            schema:
              type: string
              format: binary
      400:
        description: 请求参数错误
      404:
        description: 文件未找到
      500:
        description: 服务器内部错误
    """
    try:
        # 获取查询参数
        filepath = request.args.get('filepath')
        filename = request.args.get('filename', 'output.xlsx')
        dxf_filename = request.args.get('dxf_filename')
        
        # 检查必要参数
        if not filepath:
            return jsonify({"error": "必须提供filepath参数"}), 400
            
        # 检查文件是否存在
        if not os.path.exists(filepath):
            return jsonify({"error": f"文件未找到: {filepath}"}), 404
            
        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        # 读取JSON文件
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
        except json.JSONDecodeError as e:
            return jsonify({"error": f"JSON文件格式错误: {str(e)}"}), 400
        except Exception as e:
            return jsonify({"error": f"读取文件时出错: {str(e)}"}), 500
            
        # 调用转换函数
        file_path = save_json_to_excel(json_data, filename, dxf_filename)
        
        # 返回文件下载
        return send_file(file_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"处理请求时出错: {str(e)}")
        return jsonify({"error": f"处理过程中出错: {str(e)}"}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """
    健康检查接口
    """
    return jsonify({"status": "healthy", "service": "json_to_excel_service"})

def calculate_steel_weights_and_lengths(json_data):
    """
    计算钢材的单重、总重或总长度
    
    对于钢板类型：
    - 单重 = 长度 * 宽度 * 厚度 * 0.00000785
    - 总重 = 单重 * 数量
    
    对于角钢类型：
    - 总长 = 长度 * 数量
    
    Args:
        json_data (list/dict): 包含钢材信息的JSON数据
        
    Returns:
        list/dict: 添加计算结果后的数据
    """
    try:
        # 处理输入数据
        if isinstance(json_data, str):
            json_data = json.loads(json_data)
            
        # 创建数据副本以避免修改原始数据
        processed_data = json_data
        
        # 如果是字典格式，检查是否包含列表数据
        if isinstance(processed_data, dict):
            # 遍历字典中的每个键值对
            for key, value in processed_data.items():
                if isinstance(value, list):
                    # 处理列表中的每个项目
                    for item in value:
                        if isinstance(item, dict):
                            _calculate_item_weight_or_length(item)
                elif isinstance(value, dict):
                    _calculate_item_weight_or_length(value)
                    
        elif isinstance(processed_data, list):
            # 处理列表中的每个项目
            for item in processed_data:
                if isinstance(item, dict):
                    _calculate_item_weight_or_length(item)
                    
        return processed_data
        
    except Exception as e:
        logger.error(f"计算钢材重量和长度时出错: {str(e)}")
        raise e

def _calculate_item_weight_or_length(item):
    """
    计算单个钢材项目的重量或长度
    
    Args:
        item (dict): 单个钢材项目的数据
    """
    try:
        steel_type = item.get('类型') or item.get('type') or item.get('steel_type')
        
        if steel_type == '钢板':
            # 获取必要参数
            length = float(item.get('长度') or item.get('length') or 0)
            width = float(item.get('宽度') or item.get('width') or 0)
            thickness = float(item.get('厚度') or item.get('thickness') or 0)
            quantity = float(item.get('数量') or item.get('quantity') or 0)
            
            # 计算单重和总重
            unit_weight = length * width * thickness * 0.00000785
            total_weight = unit_weight * quantity
            
            # 添加计算结果到项目中
            item['单重'] = round(unit_weight, 1)
            item['总重'] = round(total_weight, 1)
            
        elif steel_type == '角钢':
            # 获取必要参数
            length = float(item.get('长度') or item.get('length') or 0)
            quantity = float(item.get('数量') or item.get('quantity') or 0)
            
            # 计算总长
            total_length = length * quantity
            
            # 添加计算结果到项目中
            item['总长'] = round(total_length, 1)
            
    except Exception as e:
        logger.warning(f"计算单项钢材数据时出错: {str(e)}, 项目数据: {item}")
        # 不中断整个处理过程，继续处理其他项目

@app.route('/calculate_steel_data', methods=['POST'])
def calculate_steel_data_endpoint():
    """
    计算钢材数据（单重、总重、总长）并返回结果
    
    ---
    parameters:
      - name: data
        in: body
        required: true
        description: 包含钢材信息的JSON数据
    responses:
      200:
        description: 成功计算并返回结果
      400:
        description: 请求参数错误
      500:
        description: 服务器内部错误
    """
    try:
        # 获取请求体中的JSON数据
        if request.is_json:
            json_data = request.get_json()
        else:
            return jsonify({"error": "请求必须包含JSON数据"}), 400
            
        # 计算钢材重量和长度
        calculated_data = calculate_steel_weights_and_lengths(json_data)
        
        # 返回计算结果
        return jsonify({
            "status": "success",
            "data": calculated_data
        })
        
    except Exception as e:
        logger.error(f"处理钢材计算请求时出错: {str(e)}")
        return jsonify({"error": f"处理过程中出错: {str(e)}"}), 500

@app.route('/calculate_and_save_steel_data', methods=['POST'])
def calculate_and_save_steel_data():
    """
    计算钢材数据并保存为Excel文件
    
    ---
    parameters:
      - name: data
        in: body
        required: true
        description: 包含钢材信息的JSON数据
      - name: filename
        in: query
        type: string
        required: false
        default: steel_data.xlsx
        description: 输出的Excel文件名
      - name: dxf_filename
        in: query
        type: string
        required: false
        description: DXF文件名，将填入Excel前三行合并单元格
    responses:
      200:
        description: 成功保存Excel文件
      400:
        description: 请求参数错误
      500:
        description: 服务器内部错误
    """
    try:
        # 获取请求体中的JSON数据
        if request.is_json:
            request_data = request.get_json()
        else:
            return jsonify({"error": "请求必须包含JSON数据"}), 400
            
        # 获取查询参数中的文件名和DXF文件名
        filename = request.args.get('filename')
        dxf_filename = request.args.get('dxf_filename')
        if not filename:
            filename = 'steel_data.xlsx'
        
        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        # 计算钢材重量和长度
        calculated_data = calculate_steel_weights_and_lengths(request_data)
        
        # 保存为Excel文件
        file_path = save_json_to_excel(calculated_data, filename, dxf_filename)
        
        # 返回文件下载
        return send_file(file_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"处理钢材计算和保存请求时出错: {str(e)}")
        return jsonify({"error": f"处理过程中出错: {str(e)}"}), 500

@app.route('/routes')
def show_routes():
    routes = []
    for rule in app.url_map.iter_rules():
        # 获取端点函数的docstring
        endpoint_func = app.view_functions.get(rule.endpoint)
        docstring = endpoint_func.__doc__.strip() if endpoint_func and endpoint_func.__doc__ else "无描述"
        
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule),
            'description': docstring
        })
    return {'routes': routes}

import argparse

# 解析命令行参数
parser = argparse.ArgumentParser()
parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
args = parser.parse_args()

# 使用指定的端口运行 Flask 应用
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=args.port, debug=True)