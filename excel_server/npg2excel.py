# npg2excel.py
# vlm_png_to_excel.py
import json
import os
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import tempfile
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import sys
import tkinter as tk
from tkinter import filedialog
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from llm_server.llm_server import VLMService

app = Flask(__name__)
vlm_service = VLMService()

class ImageToExcelConverter:
    def __init__(self):
        # 定义提取字段配置，通过修改这个列表可以同时改变提示语句和JSON示例
        self.extraction_fields = [
            {"key": "drawing_number", "label": "ABC123", "display_name": "图号"},
            {"key": "length", "label": "1200", "display_name": "长度(mm)"},
            {"key": "width", "label": "800", "display_name": "宽度(mm)"},
            {"key": "thickness", "label": "10", "display_name": "厚度(mm)"},
            {"key": "quantity", "label": "5", "display_name": "数量"},
            {"key": "material", "label": "不锈钢", "display_name": "材料"},
            {"key": "specification", "label": "SUS304", "display_name": "规格"},
            {"key": "process", "label": "切割", "display_name": "工艺"},
            {"key": "remark", "label": "不锈钢板", "display_name": "备注"}
        ]
    
    def parse_vlm_response(self, response_data):
        """
        解析 VLM 返回的结构化数据
        """
        try:
            result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            if result.startswith("```json"):
                result = result[7:]
            if result.startswith("```"):
                result = result[3:]
            if result.endswith("```"):
                result = result[:-3]
            
            result = result.strip()
            parsed = json.loads(result)
            
            if "drawing_info" in parsed:
                return {"drawing_info": parsed["drawing_info"]}
            else:
                return {}
        except Exception as e:
            print(f"解析失败: {e}")
            return {}

    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value

    def save_data_to_excel(self, data, output_path="output.xlsx", image_filename=None):
        """
        将结构化数据保存为Excel文件
        
        Args:
            data (dict): 结构化数据
            output_path (str): 输出Excel文件路径
            image_filename (str): 原始图像文件名，用于填入Excel前三行合并单元格
        """
        wb = Workbook()
        ws = wb.active
        
        if image_filename:
            ws["A1"] = image_filename
            ws.merge_cells("A1:E3")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(size=14, bold=True)
            # 设置标题行行高为25
            for row_num in range(1, 4):
                ws.row_dimensions[row_num].height = 25
            start_row = 4
        else:
            start_row = 1
        
        rows = []
        row_count = 0  # 记录数据行数
        
        if "drawing_info" in data and data["drawing_info"]:
            rows.append(["图纸信息", "", "", "", ""])  # 标题行
            row_count += 1
            
            # 动态提取所有字段名作为表头
            all_fields = set()
            for item in data["drawing_info"]:
                all_fields.update(item.keys())
            
            # 创建显示名称映射
            display_names = {field["key"]: field["display_name"] for field in self.extraction_fields}
            
            # 创建表头行（中文）
            header_row = []
            field_order = []  # 保持列顺序
            
            # 按预定义顺序添加字段
            for field_config in self.extraction_fields:
                field_key = field_config["key"]
                if field_key in all_fields:
                    header_row.append(display_names.get(field_key, field_key))
                    field_order.append(field_key)
            
            # 添加其他未预定义的字段
            for field in all_fields:
                if field not in field_order:
                    header_row.append(field)  # 对于未预定义字段，直接使用字段名
                    field_order.append(field)
                    
            rows.append(header_row)
            row_count += 1
            
            # 填充数据行
            for item in data["drawing_info"]:
                data_row = []
                for field in field_order:
                    cell_value = item.get(field, "")
                    # 尝试将值转换为数字
                    final_value = self._convert_to_number_if_possible(cell_value)
                    data_row.append(final_value)
                rows.append(data_row)
                row_count += 1
        
        # 写入Excel
        for i, row in enumerate(rows, start=start_row):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
            # 设置每行行高为25
            ws.row_dimensions[i].height = 25
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大宽度
        
        wb.save(output_path)
        print(f"Excel 文件已保存至: {output_path}")
        return output_path
    def select_image_file(self):
        """
        弹出文件选择窗口选择图像文件
        """
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_types = [
            ("Image files", "*.png *.jpg *.jpeg *.bmp *.tiff"),
            ("PNG files", "*.png"),
            ("JPEG files", "*.jpg *.jpeg"),
            ("All files", "*.*")
        ]
        
        image_path = filedialog.askopenfilename(
            title="请选择图像文件",
            filetypes=file_types
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return image_path

    def analyze_image_and_generate_excel(self, image_path):
        """
        主函数：调用 VLM 分析图像，并生成 Excel 文件
        
        Args:
            image_path (str): 输入图像路径
        """
        output_path = r"excel_output/npg2excel/extracted_data.xlsx"
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # 根据当前字段配置动态生成提示语句
        labels = [field["display_name"] for field in self.extraction_fields]
        labels_str = "、".join(labels)
        instruction = f"请提取这张图片的信息，尽可能提取出{labels_str}等信息"
        
        # 动态构建JSON示例中的字段部分
        field_lines = []
        for field in self.extraction_fields:
            field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
        
        json_fields_block = ",\n".join(field_lines)
        
        messages = [
            {
                "role": "user",
                "content": f'''{instruction}，并以 JSON 格式返回。
输出格式要求：
{{
    "drawing_info": [
        {{
{json_fields_block}
        }}
    ]
}}'''
            }
        ]

        try:
            result = vlm_service.create_with_image(messages, image_path)
            print("VLM 返回:", result)
            extracted_data = self.parse_vlm_response(result)
            self.save_data_to_excel(extracted_data, output_path, os.path.basename(image_path))
        except Exception as e:
            print("错误:", str(e))

# 创建全局转换器实例
converter = ImageToExcelConverter()

@app.route('/api/analyze-image', methods=['GET'])
def analyze_image_api():
    """
    Flask 接口：通过URL参数指定图片路径，调用VLM分析并生成Excel文件
    
    参数:
    - image_path: 图片路径 (必需)
    """
    try:
        image_path = request.args.get('image_path')
     
        if not image_path:
            return jsonify({"error": "缺少 image_path 参数"}), 400
        
        if not os.path.exists(image_path):
            return jsonify({"error": f"图像文件 '{image_path}' 不存在"}), 400
            
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
            output_path = tmp_excel.name
        
        converter.analyze_image_and_generate_excel(image_path)
        
        return send_file(
            output_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/routes')
def show_routes():
    routes = []
    for rule in app.url_map.iter_rules():
        endpoint_func = app.view_functions.get(rule.endpoint)
        docstring = endpoint_func.__doc__.strip() if endpoint_func and endpoint_func.__doc__ else "无描述"
        
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule),
            'description': docstring
        })
    return {'routes': routes}

def main():
    if len(sys.argv) == 1:
        print("请选择操作:")
        print("1. 通过文件选择对话框选择图像")
        print("2. 手动输入图像路径")
        print("3. 启动Flask服务器")
        
        choice = input("请输入选项 (1/2/3): ").strip()
        
        if choice == "1":
            # 使用文件选择对话框
            image_path = converter.select_image_file()
            
            if not image_path:
                print("未选择文件，程序退出")
                sys.exit(0)
                
            if not os.path.exists(image_path):
                print(f"错误: 图像文件 '{image_path}' 不存在")
                sys.exit(1)
                
            try:
                converter.analyze_image_and_generate_excel(image_path)
                print("处理完成！")
            except Exception as e:
                print(f"处理过程中出现错误: {e}")
                
        elif choice == "2":
            # 原有的手动输入方式
            print("请输入图像文件路径（输入 'quit' 退出）:")
            while True:
                image_path = input("图像路径: ").strip()
                
                if image_path.lower() == 'quit':
                    print("程序退出")
                    break
                    
                if not os.path.exists(image_path):
                    print(f"错误: 图像文件 '{image_path}' 不存在，请重新输入")
                    continue
                    
                try:
                    converter.analyze_image_and_generate_excel(image_path)
                    print("处理完成！")
                    break
                except Exception as e:
                    print(f"处理过程中出现错误: {e}")
                    continue
        elif choice == "3":
            # 启动Flask服务器
            import argparse
            parser = argparse.ArgumentParser()
            parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
            args = parser.parse_args()
            app.run(host='0.0.0.0', port=args.port, debug=True)
        else:
            print("无效选项，程序退出")
    else:
        import argparse
        parser = argparse.ArgumentParser()
        parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
        args = parser.parse_args()
        app.run(host='0.0.0.0', port=args.port, debug=True)

if __name__ == '__main__':
    main()