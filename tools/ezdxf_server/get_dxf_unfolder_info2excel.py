import os
import re
import pandas as pd
from ezdxf import readfile
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_thickness_from_filename(filename):
    """
    从文件名中提取板厚信息
    
    Args:
        filename (str): DXF文件名
        
    Returns:
        float: 板厚值(mm)，如果未找到则返回None
    """
    # 匹配类似 (2mm-FP1497x2468-3D D788x1497) 的模式
    pattern = r'(\d+(?:\.\d+)?)mm'
    match = re.search(pattern, filename)
    
    if match:
        return float(match.group(1))
    return None

def explode_all_blocks(msp):
    """
    递归分解模型空间中的所有块引用和多段线
    这是从base_server.py中复制的方法
    
    Args:
        msp: 模型空间对象
        
    Returns:
        tuple: (分解的块数量, 分解出的实体数量)
        
    Raises:
        ValueError: 当遇到不支持的实体类型(如REGION)时抛出异常
    """
    blocks_broken = 0
    exploded_entities = 0
    
    # 检查是否存在不支持的实体类型(如REGION)
    regions = [entity for entity in msp if entity.dxftype() == 'REGION']
    if regions:
        raise ValueError(f"不支持处理REGION实体，发现 {len(regions)} 个REGION实体")
    
    # 多次遍历直到没有更多的INSERT实体和可分解的多段线
    while True:
        # 收集所有需要分解的实体（INSERT和多段线）
        inserts = [entity for entity in msp if entity.dxftype() == 'INSERT']
        polylines = [entity for entity in msp if entity.dxftype() in ['POLYLINE', 'LWPOLYLINE']]
        
        # 如果没有需要分解的实体，则退出循环
        if not inserts and not polylines:
            break
            
        # 分解所有块引用
        for insert in inserts:
            try:
                exploded = insert.explode()
                blocks_broken += 1
                exploded_entities += len(exploded)
                logger.info(f"分解块 '{insert.dxf.name}'，获得 {len(exploded)} 个实体")
            except Exception as e:
                logger.info(f"分解块 '{insert.dxf.name}' 时出错: {e}")
                
        # 分解所有多段线
        for polyline in polylines:
            try:
                exploded = polyline.explode()
                blocks_broken += 1
                exploded_entities += len(exploded)
                logger.info(f"分解多段线，获得 {len(exploded)} 个实体")
            except Exception as e:
                logger.info(f"分解多段线时出错: {e}")
                
    return blocks_broken, exploded_entities

def calculate_dxf_extents(msp):
    """
    计算模型空间中所有实体的边界框（参考base_server.py的方法）
    
    Args:
        msp: 模型空间对象
        
    Returns:
        tuple: (min_x, min_y, max_x, max_y) 或 None(如果没有实体)
    """
    all_x_coords = []
    all_y_coords = []
    
    for entity in msp:
        try:
            if entity.dxftype() == 'LINE':
                start = entity.dxf.start
                end = entity.dxf.end
                all_x_coords.extend([start.x, end.x])
                all_y_coords.extend([start.y, end.y])
                
            elif entity.dxftype() == 'CIRCLE':
                center = entity.dxf.center
                radius = entity.dxf.radius
                all_x_coords.extend([center.x - radius, center.x + radius])
                all_y_coords.extend([center.y - radius, center.y + radius])
                
            elif entity.dxftype() == 'ARC':
                center = entity.dxf.center
                radius = entity.dxf.radius
                all_x_coords.extend([center.x - radius, center.x + radius])
                all_y_coords.extend([center.y - radius, center.y + radius])
                
            elif entity.dxftype() in ['TEXT', 'MTEXT']:
                insert = entity.dxf.insert
                all_x_coords.append(insert.x)
                all_y_coords.append(insert.y)
                
            elif entity.dxftype() == 'SPLINE':
                if hasattr(entity, 'fit_points') and entity.fit_points:
                    points = [(p.x, p.y) for p in entity.fit_points]
                elif hasattr(entity, 'control_points'):
                    points = [(p[0], p[1]) for p in entity.control_points]
                else:
                    continue
                    
                x_coords = [p[0] for p in points]
                y_coords = [p[1] for p in points]
                all_x_coords.extend(x_coords)
                all_y_coords.extend(y_coords)
                
        except Exception as e:
            logger.debug(f"计算实体 {entity.dxftype()} 边界时出错: {e}")
            continue
    
    if not all_x_coords or not all_y_coords:
        return None
        
    return (min(all_x_coords), min(all_y_coords), max(all_x_coords), max(all_y_coords))

def get_modelspace_bbox(dxf_doc):
    """
    获取模型空间的边界框
    
    Args:
        dxf_doc: DXF文档对象
        
    Returns:
        tuple: (min_x, min_y, max_x, max_y) 或 None（如果没有实体）
    """
    modelspace = dxf_doc.modelspace()
    
    # 先分解所有块引用和多段线
    try:
        blocks_broken, exploded_entities = explode_all_blocks(modelspace)
        logger.info(f"分解了 {blocks_broken} 个块，获得了 {exploded_entities} 个实体")
    except ValueError as e:
        logger.warning(f"分解块时遇到问题: {e}")
    
    # 注意：这里重新获取实体列表，因为explode操作会修改modelspace
    entities = list(modelspace)
    
    if not entities:
        return None
    
    # 使用与base_server.py相同的方法计算边界
    bbox = calculate_dxf_extents(modelspace)
    
    if bbox:
        logger.info(f"共有 {len(entities)} 个实体参与边界计算")
        
    return bbox

def process_dxf_files(folder_path):
    """
    处理文件夹中的所有DXF文件并导出信息到Excel
    
    Args:
        folder_path (str): DXF文件所在文件夹路径
    """
    # 创建输出文件夹
    output_folder = Path("dxf_output/get_dxf_unfolder_info")
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # 存储结果的列表
    results = []
    
    # 遍历文件夹中的所有DXF文件
    dxf_files = [f for f in Path(folder_path).iterdir() if f.suffix.lower() == '.dxf']
    
    for dxf_file in dxf_files:
        try:
            # 读取DXF文件
            doc = readfile(str(dxf_file))
            
            # 提取板厚信息
            thickness = extract_thickness_from_filename(dxf_file.name)
            
            # 获取模型边界框
            bbox = get_modelspace_bbox(doc)
            
            # 计算宽度和高度
            width = None
            height = None
            if bbox:
                min_x, min_y, max_x, max_y = bbox
                width = round(max_x - min_x)
                height = round(max_y - min_y)
                
                # 输出每个文件的尺寸信息
                print(f"{dxf_file.name}: 宽度={width}mm, 高度={height}mm")
            
            # 添加到结果列表
            results.append({
                '文件名': dxf_file.name,
                '板厚(mm)': thickness,
                '最小X坐标': round(bbox[0]) if bbox else None,
                '最小Y坐标': round(bbox[1]) if bbox else None,
                '最大X坐标': round(bbox[2]) if bbox else None,
                '最大Y坐标': round(bbox[3]) if bbox else None,
                '宽度(mm)': width,
                '高度(mm)': height
            })
            
            print(f"已处理: {dxf_file.name}")
            
        except Exception as e:
            print(f"处理文件 {dxf_file.name} 时出错: {e}")
            # 即使出错也记录文件信息
            results.append({
                '文件名': dxf_file.name,
                '板厚(mm)': extract_thickness_from_filename(dxf_file.name),
                '最小X坐标': None,
                '最小Y坐标': None,
                '最大X坐标': None,
                '最大Y坐标': None,
                '宽度(mm)': None,
                '高度(mm)': None
            })
    
    # 导出到Excel
    if results:
        df = pd.DataFrame(results)
        output_file = output_folder / "dxf_info.xlsx"
        df.to_excel(output_file, index=False)
        
        # 使用openpyxl进一步处理Excel文件
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        
        # 在第一行之前插入新行
        worksheet.insert_rows(1)
        
        # 合并第一行并填入文件夹路径
        folder_path_str = str(Path(folder_path).resolve())
        worksheet.cell(row=1, column=1, value=folder_path_str)
        
        # 合并所有列的单元格
        max_column = worksheet.max_column
        if max_column > 1:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        
        # 设置第一行的样式
        first_cell = worksheet.cell(row=1, column=1)
        first_cell.alignment = Alignment(horizontal='center', vertical='center')
        first_cell.font = Font(bold=True)
        
        # 设置行高为25
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 25
        
        # 保存修改后的Excel文件
        workbook.save(output_file)
        
        print(f"导出完成，共处理 {len(results)} 个文件，保存至: {output_file}")
    else:
        print("未找到任何DXF文件")

# 使用示例
if __name__ == "__main__":
    # 设置DXF文件所在的文件夹路径
    dxf_folder = input("请输入DXF文件所在的文件夹路径: ").strip()
    
    if os.path.exists(dxf_folder):
        process_dxf_files(dxf_folder)
    else:
        print("指定的文件夹不存在，请检查路径是否正确")