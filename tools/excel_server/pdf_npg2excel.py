import json
import os
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import tempfile
import sys
import tkinter as tk
from tkinter import filedialog
import logging 
import fitz  # PyMuPDF
import re

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from llm_server.llm_class import VLMService

vlm_service = VLMService()

class UniversalFileToExcelConverter:
    def __init__(self):
        # 定义提取字段配置，通过修改这个列表可以同时改变提示语句和JSON示例
        self.extraction_fields = [
            {"key": "drawing_number", "label": "P7764", "display_name": "图号"},
            {"key": "specification", "label": "PL厚度*宽度*长度（类似PL8*504*1000）", "display_name": "规格"},
            
            {"key": "quantity", "label": "5", "display_name": "数量"},
                {"key": "weight", "label": "1000", "display_name": "重量"},
            {"key": "material", "label": "不锈钢", "display_name": "材料"},
          
            {"key": "remark", "label": "不锈钢板", "display_name": "备注"}
        ]
    
    def convert_pdf_to_images(self, pdf_path):
        """
        将PDF文件转换为图像列表
        
        Args:
            pdf_path (str): PDF文件路径
            
        Returns:
            list: PIL图像对象列表
        """
        logger.info(f"正在转换PDF为图像: {pdf_path}")
        
        try:
            # 打开PDF文件
            pdf_document = fitz.open(pdf_path)
            images = []
            
            # 遍历每一页
            for page_num in range(len(pdf_document)):
                try:
                    # 获取页面
                    page = pdf_document[page_num]
                    
                    # 渲染页面为图像
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
                    pix = page.get_pixmap(matrix=mat)
                    
                    # 直接转换为PIL图像而不保存到磁盘
                    img_data = pix.tobytes("ppm")
                    pil_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    images.append(pil_image)
                    
                    logger.info(f"已转换PDF第 {page_num + 1} 页为图像")
                except Exception as page_error:
                    logger.error(f"处理PDF第 {page_num + 1} 页时出错: {str(page_error)}")
                    continue  # 继续处理其他页面
            
            pdf_document.close()
            return images
        except Exception as e:
            logger.error(f"打开PDF文件 {pdf_path} 时出错: {str(e)}")
            return []

    def parse_vlm_response(self, response_data):
        """
        解析 VLM 返回的结构化数据
        """
        try:
            result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            if result.startswith("```json"):
                result = result[7:]
            if result.startswith("```"):
                result = result[3:]
            if result.endswith("```"):
                result = result[:-3]
            
            result = result.strip()
            parsed = json.loads(result)
            
            if "drawing_info" in parsed:
                return {"drawing_info": parsed["drawing_info"]}
            else:
                return {}
        except Exception as e:
            print(f"解析失败: {e}")
            return {}

    def _convert_specification_format(self, value):
        """
        转换规格格式，添加PL前缀（如果没有）并将x替换为*，去除空格，保留一位小数
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 去除所有空格
            value = value.replace(' ', '')
            
            # 替换所有的"x"为"*"
            value = value.replace('x', '*').replace('X', '*')
            
            # 处理数字，保留一位小数
            # 使用正则表达式找到所有数字部分并格式化
            def format_number(match):
                try:
                    num = float(match.group())
                    # 如果是整数，不保留小数点
                    if num.is_integer():
                        return str(int(num))
                    else:
                        # 保留一位小数
                        return f"{num:.1f}"
                except:
                    return match.group()
            
            # 匹配数字（包括小数）
            value = re.sub(r'\d+\.?\d*', format_number, value)
            
            # 如果不以PL开头，则添加PL前缀
            if not value.upper().startswith('PL'):
                value = 'PL' + value
                
            return value
        
        return value

    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 处理Unicode数字字符（如①②③等）
            try:
                unicode_digits = {
                    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
                    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10
                }
                if value in unicode_digits:
                    return unicode_digits[value]
            except:
                pass
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value

    def save_data_to_excel(self, data, output_path="output.xlsx", image_filename=None):
        """
        将结构化数据保存为Excel文件
        
        Args:
            data (dict): 结构化数据
            output_path (str): 输出Excel文件路径
            image_filename (str): 原始图像文件名，用于填入Excel前三行合并单元格
        """
        # 防止文件名重复
        base_path, ext = os.path.splitext(output_path)
        counter = 1
        final_output_path = output_path
        while os.path.exists(final_output_path):
            final_output_path = f"{base_path}_{counter}{ext}"
            counter += 1
        
        wb = Workbook()
        ws = wb.active
        
        # 修改：将原来的三行标题改为一行
        if image_filename:
            ws["A1"] = image_filename
            ws.merge_cells("A1:E1")  # 修改：合并范围从A1:E3改为A1:E1
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(size=14, bold=True)
            # 修改：只设置第一行行高
            ws.row_dimensions[1].height = 25
            start_row = 2  # 修改：起始行从第4行改为第2行
        else:
            start_row = 1
        
        rows = []
        row_count = 0  # 记录数据行数
        
        if "drawing_info" in data and data["drawing_info"]:
            rows.append(["图纸信息", "", "", "", ""])  # 标题行
            row_count += 1
            
            # 动态提取所有字段名作为表头
            all_fields = set()
            for item in data["drawing_info"]:
                all_fields.update(item.keys())
            
            # 创建显示名称映射
            display_names = {field["key"]: field["display_name"] for field in self.extraction_fields}
            
            # 创建表头行（中文）
            header_row = []
            field_order = []  # 保持列顺序
            
            # 按预定义顺序添加字段
            for field_config in self.extraction_fields:
                field_key = field_config["key"]
                if field_key in all_fields:
                    header_row.append(display_names.get(field_key, field_key))
                    field_order.append(field_key)
            
            # 添加其他未预定义的字段
            for field in all_fields:
                if field not in field_order:
                    header_row.append(field)  # 对于未预定义字段，直接使用字段名
                    field_order.append(field)
                    
            rows.append(header_row)
            row_count += 1
            
            # 填充数据行
            for item in data["drawing_info"]:
                data_row = []
                for field in field_order:
                    cell_value = item.get(field, "")
                    # 特殊处理 specification 字段
                    if field == "specification":
                        cell_value = self._convert_specification_format(cell_value)
                    # 尝试将值转换为数字
                    final_value = self._convert_to_number_if_possible(cell_value)
                    data_row.append(final_value)
                rows.append(data_row)
                row_count += 1
        
        # 写入Excel
        for i, row in enumerate(rows, start=start_row):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
            # 设置每行行高为25
            ws.row_dimensions[i].height = 25
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大宽度
        
        wb.save(final_output_path)
        print(f"Excel 文件已保存至: {final_output_path}")
        return final_output_path   
    def select_files(self):
        """
        弹出文件选择窗口选择多个图像或PDF文件
        """
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_types = [
            ("All supported files", "*.png *.jpg *.jpeg *.bmp *.tiff *.pdf"),
            ("Image files", "*.png *.jpg *.jpeg *.bmp *.tiff"),
            ("PDF files", "*.pdf"),
            ("All files", "*.*")
        ]
        
        file_paths = filedialog.askopenfilenames(
            title="请选择一个或多个图像或PDF文件",
            filetypes=file_types
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return list(file_paths)

    def analyze_file_and_generate_excel(self, file_path):
        """
        主函数：调用 VLM 分析文件，并生成 Excel 文件
        
        Args:
            file_path (str): 输入文件路径（图像或PDF）
        """
        try:
            if file_path.lower().endswith('.pdf'):
                # 处理PDF文件
                images = self.convert_pdf_to_images(file_path)
                if not images:
                    raise Exception("无法从PDF文件生成图像")
                
                # 获取文件目录和文件名（不含扩展名）
                file_dir = os.path.dirname(file_path)
                file_name = os.path.splitext(os.path.basename(file_path))[0]
                
                # 为PDF中的每一页生成单独的Excel文件
                for page_num, image in enumerate(images):
                    # 生成输出文件路径
                    if len(images) > 1:
                        output_path = os.path.join(file_dir, f"{file_name}_page_{page_num + 1}.xlsx")
                        image_filename = f"{os.path.basename(file_path)} - 第{page_num + 1}页"
                    else:
                        output_path = os.path.join(file_dir, f"{file_name}.xlsx")
                        image_filename = os.path.basename(file_path)
                    
                    # 保存当前页为临时文件以供VLM服务使用
                    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                        image.save(tmp_file, format='PNG')
                        temp_image_path = tmp_file.name
                    
                    try:
                        # 根据当前字段配置动态生成提示语句
                        labels = [field["display_name"] for field in self.extraction_fields]
                        labels_str = "、".join(labels)
                        instruction = f"请提取这张图片的信息，尽可能提取出{labels_str}等信息"
                        
                        # 动态构建JSON示例中的字段部分
                        field_lines = []
                        for field in self.extraction_fields:
                            field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
                        
                        json_fields_block = ",\n".join(field_lines)
                        
                        # 修改：添加文件名信息到提示内容中
                        messages = [
                            {
                                "role": "user",
                                "content": f'''文件名: {os.path.basename(file_path)}
{instruction}，并以 JSON 格式返回。
输出格式要求：
{{
    "drawing_info": [
        {{
{json_fields_block}
        }}
    ]
}}'''
                            }
                        ]
                        
                        result = vlm_service.create_with_image(messages, temp_image_path)
                        print("VLM 返回:", result)
                        extracted_data = self.parse_vlm_response(result)
                        self.save_data_to_excel(extracted_data, output_path, image_filename)
                    finally:
                        # 删除临时文件
                        os.unlink(temp_image_path)
            else:
                # 处理图像文件
                file_dir = os.path.dirname(file_path)
                file_name = os.path.splitext(os.path.basename(file_path))[0]
                output_path = os.path.join(file_dir, f"{file_name}.xlsx")
                
                # 根据当前字段配置动态生成提示语句
                labels = [field["display_name"] for field in self.extraction_fields]
                labels_str = "、".join(labels)
                instruction = f"请提取这张图片的信息，尽可能提取出{labels_str}等信息"
                
                # 动态构建JSON示例中的字段部分
                field_lines = []
                for field in self.extraction_fields:
                    field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
                
                json_fields_block = ",\n".join(field_lines)
                
                # 修改：添加文件名信息到提示内容中
                messages = [
                    {
                        "role": "user",
                        "content": f'''文件名: {os.path.basename(file_path)}
{instruction}，并以 JSON 格式返回。
输出格式要求：
{{
    "drawing_info": [
        {{
{json_fields_block}
        }}
    ]
}}'''
                    }
                ]
                
                result = vlm_service.create_with_image(messages, file_path)
                print("VLM 返回:", result)
                extracted_data = self.parse_vlm_response(result)
                self.save_data_to_excel(extracted_data, output_path, os.path.basename(file_path))
        except Exception as e:
            print("错误:", str(e))

    def batch_process_files(self, file_paths):
        """
        批量处理多个文件，每个文件生成单独的Excel文件
        
        Args:
            file_paths (list): 文件路径列表
        """
        if not file_paths:
            print("未选择任何文件")
            return
        
        for i, file_path in enumerate(file_paths, 1):
            print(f"\n正在处理第 {i}/{len(file_paths)} 个文件: {file_path}")
            
            if not os.path.exists(file_path):
                print(f"警告: 文件 '{file_path}' 不存在，跳过处理")
                continue
                
            try:
                self.analyze_file_and_generate_excel(file_path)
                print(f"完成处理: {file_path}")
            except Exception as e:
                print(f"处理文件 '{file_path}' 时出现错误: {e}")

# 创建全局转换器实例
converter = UniversalFileToExcelConverter()

def main():
    # 直接选择多个文件进行批量处理
    file_paths = converter.select_files()
    
    if not file_paths:
        print("未选择文件，程序退出")
        sys.exit(0)
        
    try:
        converter.batch_process_files(file_paths)
        print("\n所有文件处理完成！")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")

if __name__ == '__main__':
    main()