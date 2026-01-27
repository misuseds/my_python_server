# organize_excel.py
import json
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import sys
import tkinter as tk
from tkinter import filedialog
import logging
import re

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# 尝试导入LLM服务类
try:
    from llm_server.llm_class import LLMService
    llm_service_available = True
except ImportError as e:
    logger.warning(f"无法导入LLM服务: {e}")
    llm_service_available = False

class ExcelProcessor:
    def __init__(self):
        if llm_service_available:
            self.llm_service = LLMService()
        else:
            self.llm_service = None
            logger.warning("LLM服务不可用")
        
        # 定义提取字段配置，通过修改这个列表可以同时改变提示语句和JSON示例
        self.extraction_fields = [
            {"key": "drawing_number", "label": "ABC123", "display_name": "图号"},
            {"key": "specification", "label": "PL厚度mm*宽度*长度", "display_name": "规格"},
            {"key": "quantity", "label": "5", "display_name": "数量"},
            {"key": "material", "label": "Q235B", "display_name": "材料"},
        ]
    
    def select_excel_files(self):
        """
        弹出文件选择窗口选择Excel文件
        """
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_types = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        file_paths = filedialog.askopenfilenames(
            title="请选择一个或多个Excel文件",
            filetypes=file_types
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return list(file_paths)
    
    def read_excel_content(self, file_path):
        """
        读取Excel文件内容并转换为文本形式
        
        Args:
            file_path (str): Excel文件路径
            
        Returns:
            dict: 包含状态和内容的字典
        """
        try:
            # 使用pandas读取Excel文件的所有工作表
            excel_data = pd.read_excel(file_path, sheet_name=None)
            
            content_text = ""
            for sheet_name, df in excel_data.items():
                content_text += f"工作表: {sheet_name}\n"
                content_text += df.to_string(index=False)
                content_text += "\n\n"
                
            return {
                'status': 'success',
                'content': content_text,
                'file_path': file_path
            }
        except Exception as e:
            logger.error(f"读取Excel文件 {file_path} 时出错: {str(e)}")
            return {
                'status': 'error',
                'message': f"读取Excel文件时出错: {str(e)}",
                'file_path': file_path
            }
    
    def parse_llm_response(self, response_data):
        """
        解析LLM返回的结构化数据（参考dxf_text_to_excel_with_llm.py的方法）
        """
        try:
            result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            # 使用正则表达式提取JSON部分
            # 匹配 ```json 或 ``` 代码块中的内容
            json_match = re.search(r'```(?:json)?\s*({.*?})\s*```', result, re.DOTALL)
            if json_match:
                result = json_match.group(1)
            else:
                # 如果没有找到代码块，尝试找到第一个 { 和最后一个 } 之间的内容
                first_brace = result.find('{')
                last_brace = result.rfind('}')
                if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
                    result = result[first_brace:last_brace+1]
            
            # 修复可能存在的换行问题：处理字符串中的意外换行
            # 1. 首先处理字符串中的换行，将被换行符分割的字符串重新连接
            # 匹配 "xxx<换行>yyy" 这样的模式并将其修复为 "xxxyyy"
            result = re.sub(r'"\s*\n\s*', '"', result)
            
            # 2. 处理可能的逗号和大括号之间的换行
            result = re.sub(r',\s*\n\s*}', '}', result)  # 处理尾随逗号后换行再跟大括号的情况
            result = re.sub(r'(\w)"\s*\n\s*"', r'\1","', result)  # 处理字段之间被换行分割的情况
            
            # 3. 移除多余的空白字符
            result = re.sub(r'\s+', ' ', result)
            
            result = result.strip()
            
            # 尝试解析JSON
            parsed = json.loads(result)
            
            if "processed_data" in parsed:
                # 对每个字段值进行清理，移除意外的换行符
                for item in parsed["processed_data"]:
                    for key in item:
                        if isinstance(item[key], str):
                            # 移除字符串中的换行符和多余空格
                            item[key] = item[key].replace('\n', '').replace('\r', '').strip()
                return {"processed_data": parsed["processed_data"]}
            else:
                return {}
        except Exception as e:
            print(f"解析失败: {e}")
            # 调试信息，帮助排查问题
            print(f"LLM返回的原始内容: {result}")
            return {}
    
    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 处理Unicode数字字符（如①②③等）
            try:
                unicode_digits = {
                    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
                    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10
                }
                if value in unicode_digits:
                    return unicode_digits[value]
            except:
                pass
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value

    def save_processed_data_to_excel(self, data, original_file_path):
        """
        将处理后的数据保存为新的Excel文件
        
        Args:
            data (dict): 处理后的结构化数据
            original_file_path (str): 原始文件路径
        """
        # 生成新文件名，在原文件名后加上"_processed"
        file_dir = os.path.dirname(original_file_path)
        file_name = os.path.splitext(os.path.basename(original_file_path))[0]
        file_ext = os.path.splitext(original_file_path)[1]
        new_file_path = os.path.join(file_dir, f"{file_name}_processed{file_ext}")
        
        # 防止文件名重复
        base_path, ext = os.path.splitext(new_file_path)
        counter = 1
        final_output_path = new_file_path
        while os.path.exists(final_output_path):
            final_output_path = f"{base_path}_{counter}{ext}"
            counter += 1
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Data"
        
        # 设置表头
        # 创建显示名称映射
        display_names = {field["key"]: field["display_name"] for field in self.extraction_fields}
        
        # 创建表头行（中文）
        header_row = []
        field_order = []  # 保持列顺序
        
        # 按预定义顺序添加字段
        for field_config in self.extraction_fields:
            field_key = field_config["key"]
            header_row.append(display_names.get(field_key, field_key))
            field_order.append(field_key)
            
        # 写入表头
        for j, header in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=j, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        if "processed_data" in data and data["processed_data"]:
            for row_idx, item in enumerate(data["processed_data"], 2):
                for col_idx, field in enumerate(field_order):
                    cell_value = item.get(field, "")
                    # 尝试将值转换为数字
                    final_value = self._convert_to_number_if_possible(cell_value)
                    ws.cell(row=row_idx, column=col_idx+1, value=final_value)
                
                # 设置行高
                ws.row_dimensions[row_idx].height = 25
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存文件
        wb.save(final_output_path)
        print(f"处理后的Excel文件已保存至: {final_output_path}")
        return final_output_path
    
    def process_excel_file(self, file_path):
        """
        主函数：读取Excel文件，调用LLM处理，并生成新的Excel文件
        
        Args:
            file_path (str): 输入Excel文件路径
        """
        try:
            # 检查LLM服务是否可用
            if not self.llm_service:
                raise Exception("LLM服务不可用")
            
            # 读取Excel内容
            excel_data = self.read_excel_content(file_path)
            if excel_data['status'] != 'success':
                raise Exception(excel_data['message'])
            
            # 根据当前字段配置动态生成提示语句
            labels = [field["display_name"] for field in self.extraction_fields]
            labels_str = "、".join(labels)
            instruction = f"请分析以下Excel表格内容，提取出{labels_str}等关键信息，并以JSON格式返回。"
            
            # 动态构建JSON示例中的字段部分
            field_lines = []
            for field in self.extraction_fields:
                field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
            
            json_fields_block = ",\n".join(field_lines)
            
            # 构建提示消息（采用dxf_text_to_excel_with_llm.py的结构化提示方式）
            messages = [
                {
                    "role": "user",
                    "content": f'''{instruction}
Excel内容：
{excel_data['content']}

输出格式要求：
{{
    "processed_data": [
        {{
{json_fields_block}
        }}
    ]
}}

注意事项：
1. 只提取包含完整信息的条目
2. 保持原有单位（如有）
3. 如有计算公式需求，请在相应字段中体现
4. 确保返回有效的JSON格式，只返回JSON代码块
5. 请确保字符串值在一行内，不要换行
6. 请确保JSON格式正确，不要在字符串中包含换行符
7. 请确保所有字段都有对应的值，不要遗漏任何字段'''
                }
            ]
            
            # 调用LLM服务
            logger.info("正在使用LLM分析Excel内容")
            result = self.llm_service.create(messages)
            logger.info("LLM分析完成")
            print("LLM 返回:", result)
            
            # 解析LLM返回结果
            processed_data = self.parse_llm_response(result)
            
            # 保存处理后的数据到新的Excel文件
            self.save_processed_data_to_excel(processed_data, file_path)
            
        except Exception as e:
            print(f"处理文件 '{file_path}' 时出现错误: {e}")
    
    def batch_process_excel_files(self, file_paths):
        """
        批量处理多个Excel文件
        
        Args:
            file_paths (list): Excel文件路径列表
        """
        if not file_paths:
            print("未选择任何文件")
            return
        
        for i, file_path in enumerate(file_paths, 1):
            print(f"\n正在处理第 {i}/{len(file_paths)} 个文件: {file_path}")
            
            if not os.path.exists(file_path):
                print(f"警告: 文件 '{file_path}' 不存在，跳过处理")
                continue
                
            try:
                self.process_excel_file(file_path)
                print(f"完成处理: {file_path}")
            except Exception as e:
                print(f"处理文件 '{file_path}' 时出现错误: {e}")

def main():
    # 创建处理器实例
    processor = ExcelProcessor()
    
    # 选择Excel文件进行批量处理
    file_paths = processor.select_excel_files()
    
    if not file_paths:
        print("未选择文件，程序退出")
        sys.exit(0)
        
    try:
        processor.batch_process_excel_files(file_paths)
        print("\n所有Excel文件处理完成！")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")

if __name__ == '__main__':
    main()