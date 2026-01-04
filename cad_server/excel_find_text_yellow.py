import re
from openpyxl import load_workbook
from pyautocad import Autocad, APoint
from dotenv import load_dotenv
import os
from _ctypes import COMError
import tkinter as tk
from tkinter import filedialog

def read_excel_data(file_path, sheet_name, column_letter):
    """
    读取Excel指定sheet和列的数据（优化版）
    
    Args:
        file_path (str): Excel文件路径
        sheet_name (str): 工作表名称
        column_letter (str): 列字母标识(如'A', 'B'等)
      
    Returns:
        list: 该列的所有数据
    """
    # 使用只读模式和data_only选项提升性能
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    
    data_list = []
    # 使用iter_rows提高效率
    for row in sheet.iter_rows(min_col=ord(column_letter.upper())-ord('A')+1, 
                              max_col=ord(column_letter.upper())-ord('A')+1,
                              values_only=True):
        cell_value = row[0]
        if cell_value is not None:
            data_list.append(str(cell_value))
    
    workbook.close()  # 显式关闭工作簿
    return data_list

def process_autocad_text(text):
    """
    处理AutoCAD文本中的特殊字符
    
    Args:
        text (str): 原始AutoCAD文本
        
    Returns:
        str: 处理后的文本
    """
    # 处理常见的AutoCAD控制字符
    replacements = {
        '\\P': ' ',      # 换行符
        '\\~': ' ',      # 非断行空格
        '\\-': '-',      # 可断行连字符
        '\\ ': ' ',      # 转义空格
    }
    
    processed_text = text
    for old, new in replacements.items():
        processed_text = processed_text.replace(old, new)
    
    # 标准化空格
    return ' '.join(processed_text.split())

def match_numbers_and_english_with_dash(text):  
    """  
    匹配文本中同时包含数字和连字符(-)的文本  
      
    Args:  
        text (str): 输入文本  
      
    Returns:  
        list: 匹配到的包含数字和连字符的字符串列表  
    """  
    # 清理文本，标准化空格
    cleaned_text = ' '.join(text.split())
    
    # 更严格的正则表达式，确保匹配项至少包含一个数字和一个连字符，且总长度至少为2
    pattern = r'(?=.*\d)(?=.*-)[0-9a-zA-Z\-]{2,}'
    matches = re.findall(pattern, cleaned_text)  
    
    # 过滤掉无效的匹配项
    filtered_matches = [match for match in matches 
                       if any(c.isdigit() for c in match) and '-' in match and match != '-']
    
    return filtered_matches

def is_text_match(auto_text, match_text):
    """
    判断AutoCAD文本是否与匹配文本相符
    
    Args:
        auto_text (str): 处理后的AutoCAD文本
        match_text (str): Excel中提取的匹配文本
        
    Returns:
        bool: 是否匹配
    """
    auto_text_lower = auto_text.lower()
    match_text_lower = match_text.lower()
    
    # 直接包含检查
    if match_text_lower in auto_text_lower:
        return True
    
    # 如果匹配文本以连字符结尾，尝试去掉连字符进行匹配
    if match_text_lower.endswith('-'):
        trimmed_match = match_text_lower[:-1]
        if trimmed_match in auto_text_lower:
            # 确保匹配的位置后面不是字母或数字，避免误匹配
            match_pos = auto_text_lower.find(trimmed_match)
            if match_pos != -1:
                # 检查匹配位置后面是否是分隔符（空格、连字符等）或结束
                end_pos = match_pos + len(trimmed_match)
                if end_pos >= len(auto_text_lower) or \
                   not auto_text_lower[end_pos].isalnum():
                    return True
    
    return False

def ensure_layer_exists(acad, layer_name):
    """
    确保图层存在，如果不存在则创建
    
    Args:
        acad: AutoCAD应用程序对象
        layer_name (str): 图层名称
        
    Returns:
        layer_obj: 图层对象
    """
    try:
        layer_obj = acad.doc.Layers.Item(layer_name)
        print(f"✓ 图层 '{layer_name}' 已存在，无需创建。")
    except:
        print(f"⚠ 图层 '{layer_name}' 不存在，正在创建...")
        try:
            layer_obj = acad.doc.Layers.Add(layer_name)
            print(f"✓ 成功创建图层: '{layer_name}'")
        except Exception as create_err:
            print(f"✗ 创建图层失败: {create_err}")
            return None
    return layer_obj

def change_autocad_text_layer(matched_texts, new_layer_name):
    """  
    在AutoCAD中查找匹配的文字并修改图层和颜色
    
    Args:  
        matched_texts (list): 需要匹配的文本列表  
        new_layer_name (str): 新图层名称（基础名称）  
    """
    acad = Autocad(create_if_not_exists=True)
    
    success_count = 0
    error_count = 0
    matched_items = []  # 用于记录实际匹配成功的文本
    error_items = []    # 用于记录处理失败的文本
    
    # 用于统计每个匹配项的匹配次数
    match_counts = {match_text: 0 for match_text in matched_texts}
    
    # 存储已创建的图层对象，避免重复创建
    created_layers = {}

    try:
        # 遍历所有文本对象
        for text_entity in acad.iter_objects(['Text', 'MText']):
            try:
                if text_entity.ObjectName not in ['AcDbText', 'AcDbMText']:
                    error_items.append(("未知类型", str(text_entity.ObjectName) if hasattr(text_entity, 'ObjectName') else "无对象类型"))
                    error_count += 1
                    continue
                # 验证对象有效性
                _ = text_entity.Handle
                
                # 获取对象类型
                obj_name = text_entity.ObjectName
                text_content = ""
                
                # 根据对象类型获取文本内容
                if obj_name == 'AcDbText':
                    text_content = text_entity.TextString
                elif obj_name == 'AcDbMText':
                    text_content = text_entity.TextString
                else:
                    error_items.append((text_content or "无法获取文本", "不支持的对象类型"))
                    error_count += 1
                    continue
                    
                # 处理AutoCAD特殊字符并标准化文本内容
                processed_text = process_autocad_text(text_content)
                normalized_text = processed_text
                print(f"正在处理{obj_name}对象: '{normalized_text}'")
                
                # 检查是否匹配
                found_matches = []  # 记录所有匹配项
                for match_text in matched_texts:
                    # 使用自定义匹配函数
                    if is_text_match(normalized_text, match_text):
                        found_matches.append(match_text)
                        match_counts[match_text] += 1  # 增加匹配计数
                
                # 添加调试信息
                if "42-5-1-2-9A" in normalized_text and not found_matches:
                    print(f"⚠ 注意: 包含'42-5-1-2-9A'的文本未匹配: '{normalized_text}'")
                    print(f"  可用匹配项: {[m for m in matched_texts if '42-5' in m][:5]}")  # 显示相关匹配项供参考
                
                # 如果有匹配项，则执行修改操作
                if found_matches:
                    # 使用第一个匹配项作为图层名称
                    target_layer = found_matches[0]
                    
                    # 确保图层存在
                    if target_layer not in created_layers:
                        layer_obj = ensure_layer_exists(acad, target_layer)
                        if layer_obj is None:
                            error_items.append((normalized_text, f"无法创建或获取图层: {target_layer}"))
                            print(f"✗ 无法创建或获取图层: {target_layer}")
                            error_count += 1
                            continue
                        created_layers[target_layer] = layer_obj
                    
                    try:
                        text_entity.Layer = target_layer
                        # 设置颜色为黄色 (颜色索引为 2)
                        text_entity.Color = 2
                        print(f"✓ 已将文字'{normalized_text}'的图层修改为'{target_layer}'，颜色设为黄色")
                        success_count += 1
                        # 记录所有匹配项
                        for match in found_matches:
                            matched_items.append((normalized_text, match))
                    except Exception as layer_error:
                        error_items.append((normalized_text, f"修改图层或颜色失败: {layer_error}"))
                        print(f"✗ 修改图层或颜色失败: {layer_error}")
                        error_count += 1
                else:
                    print(f"ℹ 文字'{normalized_text}'未找到匹配项，跳过处理")
                    
            except COMError as e:
                # COM错误,跳过无效对象
                error_items.append((str(text_entity) if hasattr(text_entity, '__str__') else "COM对象", f"COM错误: {e}"))
                error_count += 1
                continue
            except AttributeError as e:
                # 属性不存在,跳过
                error_items.append(("属性错误", str(e)))
                error_count += 1
                continue
            except Exception as e:
                # 其他错误,记录并继续
                error_items.append((str(text_entity) if hasattr(text_entity, '__str__') else "未知对象", f"其他错误: {e}"))
                error_count += 1
                continue
        
        # 打印匹配成功的项目
        print(f"\n匹配成功的项目:")
        for text_content, match_text in matched_items:
            print(f"  AutoCAD文本: '{text_content}' <-> Excel匹配项: '{match_text}'")
        
        # 打印匹配次数统计
        print(f"\n匹配项统计:")
        sorted_match_counts = sorted(match_counts.items(), key=lambda x: x[1], reverse=True)
        for match_text, count in sorted_match_counts:
            if count > 0:
                print(f"  '{match_text}': {count} 次")
        
        # 如果有错误项，则打印详细错误信息
        if error_items:
            print(f"\n处理失败的项目:")
            for i, (content, error) in enumerate(error_items, 1):
                print(f"  {i}. 内容: '{content}' -> 错误: {error}")
        
        print(f"\n处理完成: 成功修改 {success_count} 个对象, 跳过 {error_count} 个无效对象")
                
    except Exception as e:
        print(f"遍历AutoCAD对象时出错: {e}")

def add_tiaoliao_to_matched_text(matched_texts):
    """  
    在AutoCAD中查找匹配的文字，修改颜色为黄色并在其后添加"挑料"
    
    Args:  
        matched_texts (list): 需要匹配的文本列表  
    """
    acad = Autocad(create_if_not_exists=True)
    
    success_count = 0
    error_count = 0
    matched_items = []  # 用于记录实际匹配成功的文本
    error_items = []    # 用于记录处理失败的文本
    
    # 用于统计每个匹配项的匹配次数
    match_counts = {match_text: 0 for match_text in matched_texts}
    
    try:
        # 遍历所有文本对象
        for text_entity in acad.iter_objects(['Text', 'MText']):
            try:
                if text_entity.ObjectName not in ['AcDbText', 'AcDbMText']:
                    error_items.append(("未知类型", str(text_entity.ObjectName) if hasattr(text_entity, 'ObjectName') else "无对象类型"))
                    error_count += 1
                    continue
                # 验证对象有效性
                _ = text_entity.Handle
                
                # 获取对象类型
                obj_name = text_entity.ObjectName
                text_content = ""
                
                # 根据对象类型获取文本内容
                if obj_name == 'AcDbText':
                    text_content = text_entity.TextString
                elif obj_name == 'AcDbMText':
                    text_content = text_entity.TextString
                else:
                    error_items.append((text_content or "无法获取文本", "不支持的对象类型"))
                    error_count += 1
                    continue
                    
                # 处理AutoCAD特殊字符并标准化文本内容
                processed_text = process_autocad_text(text_content)
                normalized_text = processed_text
                print(f"正在处理{obj_name}对象: '{normalized_text}'")
                
                # 检查是否匹配
                found_matches = []  # 记录所有匹配项
                for match_text in matched_texts:
                    # 使用自定义匹配函数
                    if is_text_match(normalized_text, match_text):
                        found_matches.append(match_text)
                        match_counts[match_text] += 1  # 增加匹配计数
                
                # 添加调试信息
                if "42-5-1-2-9A" in normalized_text and not found_matches:
                    print(f"⚠ 注意: 包含'42-5-1-2-9A'的文本未匹配: '{normalized_text}'")
                    print(f"  可用匹配项: {[m for m in matched_texts if '42-5' in m][:5]}")  # 显示相关匹配项供参考
                
                # 如果有匹配项，则执行修改操作
                if found_matches:
                    try:
                        # 设置颜色为黄色 (颜色索引为 2)
                        text_entity.Color = 2
                        
                        # 在原有文本基础上添加"挑料"
                        if obj_name == 'AcDbText':
                            original_text = text_entity.TextString
                            # 避免重复添加"挑料"
                            if not original_text.endswith("挑料"):
                                text_entity.TextString = original_text + "挑料"
                        elif obj_name == 'AcDbMText':
                            original_text = text_entity.TextString
                            # 避免重复添加"挑料"
                            if not original_text.endswith("挑料"):
                                text_entity.TextString = original_text + "挑料"
                        
                        print(f"✓ 已将文字'{normalized_text}'的颜色设为黄色并在其后添加'挑料'")
                        success_count += 1
                        # 记录所有匹配项
                        for match in found_matches:
                            matched_items.append((normalized_text, match))
                    except Exception as modify_error:
                        error_items.append((normalized_text, f"修改文本内容或颜色失败: {modify_error}"))
                        print(f"✗ 修改文本内容或颜色失败: {modify_error}")
                        error_count += 1
                else:
                    print(f"ℹ 文字'{normalized_text}'未找到匹配项，跳过处理")
                    
            except COMError as e:
                # COM错误,跳过无效对象
                error_items.append((str(text_entity) if hasattr(text_entity, '__str__') else "COM对象", f"COM错误: {e}"))
                error_count += 1
                continue
            except AttributeError as e:
                # 属性不存在,跳过
                error_items.append(("属性错误", str(e)))
                error_count += 1
                continue
            except Exception as e:
                # 其他错误,记录并继续
                error_items.append((str(text_entity) if hasattr(text_entity, '__str__') else "未知对象", f"其他错误: {e}"))
                error_count += 1
                continue
        
        # 打印匹配成功的项目
        print(f"\n匹配成功的项目:")
        for text_content, match_text in matched_items:
            print(f"  AutoCAD文本: '{text_content}' <-> Excel匹配项: '{match_text}'")
        
        # 打印匹配次数统计
        print(f"\n匹配项统计:")
        sorted_match_counts = sorted(match_counts.items(), key=lambda x: x[1], reverse=True)
        for match_text, count in sorted_match_counts:
            if count > 0:
                print(f"  '{match_text}': {count} 次")
        
        # 如果有错误项，则打印详细错误信息
        if error_items:
            print(f"\n处理失败的项目:")
            for i, (content, error) in enumerate(error_items, 1):
                print(f"  {i}. 内容: '{content}' -> 错误: {error}")
        
        print(f"\n处理完成: 成功修改 {success_count} 个对象, 跳过 {error_count} 个无效对象")
                
    except Exception as e:
        print(f"遍历AutoCAD对象时出错: {e}")

# 同时需要修改主函数中的调用部分
def main(sheet_name, column_letter):  
    """  
    主函数:读取Excel数据并在AutoCAD中修改匹配文字（添加"挑料"）  
      
    Args:  
        sheet_name (str): 工作表名称  
        column_letter (str): 列字母标识  
    """  
    print("=" * 50)  
    print("AutoCAD文字批量添加挑料工具")  
    print("=" * 50)  
    
    # 弹窗选择Excel文件
    print("\n[0/3] 请选择Excel文件...")
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    excel_file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
    )
    root.destroy()
    
    if not excel_file_path:
        print("✗ 未选择文件，程序退出")
        return
    
    # 1. 读取Excel数据  
    print("\n[1/3] 正在读取Excel数据...")  
    try:  
        excel_data = read_excel_data(excel_file_path, sheet_name, column_letter)  
        print(f"✓ 读取到 {len(excel_data)} 条数据")  
    except Exception as e:  
        print(f"✗ 读取Excel失败: {e}")  
        return  
      
    # 2. 提取匹配的包含数字和连字符的文本  
    print("\n[2/3] 正在匹配包含数字和连字符的文本...")  
    matched_texts = []  
    for data in excel_data:  
        matches = match_numbers_and_english_with_dash(data)  
        matched_texts.extend(matches)  
    
    # 去重  
    matched_texts = list(set(matched_texts))  
    print(f"✓ 找到 {len(matched_texts)} 个匹配项")  
    if matched_texts:  
        print(f"  匹配项示例: {matched_texts[:5]}")  
    
    # 3. 在AutoCAD中给匹配文字添加"挑料"
    print(f"\n[3/3] 正在给AutoCAD中匹配文字添加'挑料'...")  
    add_tiaoliao_to_matched_text(matched_texts)  
    
    print("\n" + "=" * 50)  
    print("操作完成!")  
    print("=" * 50)  

# 更新主程序入口
if __name__ == "__main__":  
    # 配置参数  
    dotenv_path = r'E:\code\apikey\.env'  
    load_dotenv(dotenv_path)  
    
    sheet_name = "挑料"  
    column = "A"  
      
    # 执行主函数  
    main(sheet_name, column)