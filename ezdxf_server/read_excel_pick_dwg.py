import openpyxl
import os
import shutil
from flask import Flask, jsonify, request

# 初始化Flask应用
app = Flask(__name__)

def find_dwg_files_by_excel_data(excel_path, folder_path, output_folder='dxf_output/pick_dxf', column='B', sheet_name=None):
    """
    读取Excel文件指定列的内容，在文件夹中查找包含这些内容的DWG文件
    
    Args:
        excel_path (str): Excel文件路径
        folder_path (str): 要搜索的文件夹路径
        output_folder (str): 输出文件夹路径
        column (str): 要读取的列名，默认为'B'
        sheet_name (str): 工作表名称，默认为None，表示使用活动工作表
    
    Returns:
        dict: 包含匹配结果的字典，键为Excel中的内容，值为匹配的文件列表
    """
    
    # 检查文件和文件夹是否存在
    if not os.path.exists(excel_path):
        print(f"Excel文件不存在: {excel_path}")
        return {}
    
    if not os.path.exists(folder_path):
        print(f"文件夹不存在: {folder_path}")
        return {}
    
    # 创建输出文件夹（如果不存在）
    os.makedirs(output_folder, exist_ok=True)
    print(f"输出文件夹已创建或已存在: {output_folder}")
    
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        
        # 根据sheet_name参数选择工作表
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                print(f"使用工作表: {sheet_name}")
            else:
                print(f"工作表 '{sheet_name}' 不存在，可用的工作表: {workbook.sheetnames}")
                return {}
        else:
            worksheet = workbook.active
            print(f"使用默认工作表: {worksheet.title}")
        
        # 读取指定列的所有内容（从第2行开始，假设第1行为标题）
        search_terms = []
        row = 2  # 从第二行开始读取数据
        while True:
            cell_value = worksheet[f'{column}{row}'].value
            if cell_value is None:
                break
            search_terms.append(str(cell_value))
            row += 1
        
        print(f"从Excel中读取到 {len(search_terms)} 个搜索项: {search_terms}")
        
        # 在文件夹中查找包含这些内容的DWG文件
        results = {}
        dwg_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.dxf')]
        
        for term in search_terms:
            matching_files = []
            copied_files = []
            for dwg_file in dwg_files:
                # 检查文件名是否包含Excel中的内容
                if term in dwg_file:
                    source_path = os.path.join(folder_path, dwg_file)
                    dest_path = os.path.join(output_folder, dwg_file)
                    matching_files.append(source_path)
                    
                    # 复制文件到输出文件夹
                    try:
                        shutil.copy2(source_path, dest_path)
                        copied_files.append(dest_path)
                        print(f"已复制文件: {dwg_file} -> {dest_path}")
                    except Exception as e:
                        print(f"复制文件 {dwg_file} 时出错: {e}")
            
            results[term] = copied_files  # 返回复制后的文件路径
            if matching_files:
                print(f"找到包含 '{term}' 的文件: {matching_files}")
            else:
                print(f"未找到包含 '{term}' 的DWG文件")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出错: {e}")
        return {}

# Flask路由：查找DWG文件接口
@app.route('/find_dwg/<path:excel_path>/<path:folder_path>')
def find_dwg_api(excel_path, folder_path):
    """
    API接口：根据Excel文件内容查找DWG文件
    支持查询参数: column(默认为'B'), sheet_name(默认为活动工作表)
    """
    column = request.args.get('column', 'B')
    sheet_name = request.args.get('sheet_name', None)
    output_folder = request.args.get('output_folder', 'dxf_output/pick_dxf')
    
    # 这里应该使用传入的路径参数而不是硬编码的路径
    # 以下两行仅用于测试目的，实际部署时应删除或修改

    results = find_dwg_files_by_excel_data(excel_path, folder_path, output_folder, column, sheet_name)
    return jsonify(results)

# Flask路由：显示所有路由
@app.route('/routes')
def show_routes():
    """
    显示所有可用的API路由
    """
    routes = []
    for rule in app.url_map.iter_rules():
        # 获取端点函数的docstring
        endpoint_func = app.view_functions.get(rule.endpoint)
        docstring = endpoint_func.__doc__.strip() if endpoint_func and endpoint_func.__doc__ else "无描述"
        
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule),
            'description': docstring
        })
    return jsonify({'routes': routes})

import argparse

# 解析命令行参数
parser = argparse.ArgumentParser()
parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
args = parser.parse_args()

# 使用指定的端口运行 Flask 应用
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=args.port, debug=True)