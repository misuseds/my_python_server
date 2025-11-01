#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
读取DXF文件的文字信息，使用LLM分析并导出到Excel
支持DWG文件自动转换为DXF后处理
"""
import os
import json
import sys
import re
import pandas as pd
import subprocess
import time
from ezdxf import readfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 添加父目录到路径，以便导入LLM服务
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# 尝试导入LLM服务类
import traceback
try:
    # 从llm_server导入LLMService
    from llm_server.llm_server import LLMService
    llm_service_available = True
except ImportError as e:
    logger.warning(f"无法导入LLM服务: {e}")
    llm_service_available = False
    logger.debug(traceback.format_exc())

class DXFTextToExcelConverter:
    def __init__(self):
        if llm_service_available:
            self.llm_service = LLMService()
        else:
            self.llm_service = None
            logger.warning("LLM服务不可用，将使用模拟数据进行演示")
        
        # 定义提取字段配置，通过修改这个列表可以同时改变提示语句和JSON示例
        self.extraction_fields = [
            {"key": "零件号", "label": "编号101"},
            {"key": "长度", "label": "1050"},
            {"key": "数量", "label": "1"},
        ]
    
    def convert_dwg_to_dxf(self, dwg_file_path):
        """
        使用ODA File Converter将DWG文件转换为DXF文件
        
        Args:
            dwg_file_path (str): DWG文件路径
            
        Returns:
            dict: 转换结果，包含状态和DXF文件路径或错误信息
        """
        start_time = time.time()
        try:
            # 检查DWG文件是否存在
            if not os.path.exists(dwg_file_path):
                return {
                    "status": "error",
                    "message": f"DWG文件不存在: {dwg_file_path}"
                }

            # 获取文件目录和文件名
            dwg_dir = os.path.dirname(dwg_file_path)
            dwg_filename = os.path.basename(dwg_file_path)
            dxf_filename = dwg_filename.replace('.dwg', '.dxf')
            dxf_file_path = dwg_file_path.replace('.dwg', '.dxf')
            if dxf_file_path == dwg_file_path:
                dxf_file_path = f"{dwg_file_path}.dxf"

            # 创建输出目录（如果不存在）
            output_dir = os.path.dirname(dxf_file_path)
            os.makedirs(output_dir, exist_ok=True)

            # 构建命令
            cmd = [
                "ODAFileConverter",
                dwg_dir,
                output_dir,
                "ACAD2018",   # 输入版本
                "DXF",        # 输出格式
                "0",          # 非递归
                "1",          # 执行审计
                "*.DWG"       # 只处理DWG文件
            ]

            logger.info(f"执行DWG转换命令: {' '.join(cmd)}")

            # 执行转换命令
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=120)

            # 记录转换过程信息
            logger.debug("ODA 输出: %s", result.stdout)
            logger.debug("ODA 错误: %s", result.stderr)

            if result.returncode == 0:
                # 检查DXF文件是否生成
                if os.path.exists(dxf_file_path):
                    response = {
                        "status": "success",
                        "message": f"DWG文件已成功转换为DXF: {dxf_file_path}",
                        "dxf_path": dxf_file_path
                    }
                    return response
                else:
                    response = {
                        "status": "error",
                        "message": f"转换完成但未生成DXF文件。请检查输出目录: {output_dir}"
                    }
                    return response
            else:
                response = {
                    "status": "error",
                    "message": f"ODA转换失败: {result.stderr}"
                }
                return response

        except subprocess.TimeoutExpired:
            return {"status": "error", "message": "转换超时"}
        except FileNotFoundError:
            return {"status": "error", "message": "未找到ODAFileConverter命令，请确保ODA File Converter已正确安装并加入PATH"}
        except Exception as e:
            return {"status": "error", "message": f"转换过程中发生错误: {str(e)}"}
    
    def read_dxf_texts(self, dxf_file_path):
        """
        读取DXF文件中的所有文本实体，包括标注文字
        """
        logger.info(f"开始读取DXF文件: {dxf_file_path}")
        
        # 检查文件是否存在
        if not os.path.exists(dxf_file_path):
            raise FileNotFoundError(f"文件不存在: {dxf_file_path}")
        
        # 检查文件格式
        file_ext = os.path.splitext(dxf_file_path)[1].lower()
        if file_ext != '.dxf':
            raise ValueError(f"不支持的文件格式: {file_ext}。仅支持DXF文件。")
        
        try:
            # 读取DXF文件
            doc = readfile(dxf_file_path)
            msp = doc.modelspace()
            
            texts = []
            for entity in msp:
                if entity.dxftype() == 'TEXT':
                    texts.append({
                        'content': entity.dxf.text,
                        'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                    })
                elif entity.dxftype() == 'MTEXT':
                    texts.append({
                        'content': entity.text,
                        'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                    })
                elif entity.dxftype() == 'DIMENSION':
                    # 处理标注实体
                    try:
                        # 获取标注文字
                        dim_text = entity.dxf.text
                        if dim_text and dim_text.strip():
                            texts.append({
                                'content': dim_text,
                                'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                            })
                        # 如果标注文字为空，尝试获取测量值
                        elif entity.dim_text:
                            texts.append({
                                'content': entity.dim_text,
                                'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                            })
                    except Exception as e:
                        logger.debug(f"处理标注实体时出错: {e}")
                        continue
            
            all_text_content = '\n'.join([text['content'] for text in texts])
            logger.info(f"成功读取DXF文件，共提取 {len(texts)} 个文本实体")
            
            return {
                'status': 'success',
                'texts': texts,
                'all_text_content': all_text_content,
                'text_count': len(texts),
                'file_path': dxf_file_path
            }
        except Exception as e:
            logger.error(f"处理文件时出错: {str(e)}")
            raise
    
    def analyze_text_with_llm(self, text_content, file_name):
        """
        使用LLM分析文本内容
        """
        
        # 根据当前字段配置动态生成提示语句
        labels = [field["label"] for field in self.extraction_fields]
        labels_str = "、".join(labels)
        instruction = f"请分析以下DXF文件中的文本信息，提取出{labels_str}等关键信息，并以JSON格式返回。"
        
        # 动态构建JSON示例中的字段部分
        field_lines = []
        for field in self.extraction_fields:
            field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
        
        json_fields_block = ",\n".join(field_lines)
        
        logger.info("正在使用LLM分析文本内容")
        
        messages = [
            {
                "role": "user",
                "content": f'''
{instruction}
文本内容：
{text_content}

输出格式要求：
{{
    "drawing_info": [
        {{
{json_fields_block}
        }}
    ]
}}
'''
            }
        ]
        
        try:
            # 调用LLM服务进行文本分析
            result = self.llm_service.create(messages)
            logger.info(f"LLM分析完成")
            print(f"LLM原始返回: {result}")
            
            # 解析LLM返回结果
            return self.parse_llm_response(result)
        except Exception as e:
            logger.error(f"LLM分析出错: {str(e)}")
            raise
    
    def parse_llm_response(self, response_data):
        """
        解析LLM返回的结构化数据，只提取第一个有效的JSON块
        """
        try:
            if isinstance(response_data, dict):
                result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            else:
                result = str(response_data)
            
            # 处理可能的Markdown格式，只提取第一个代码块
            if "```json" in result:
                # 提取第一个```json和```之间的内容
                start = result.find("```json") + 7
                end = result.find("```", start)
                if end == -1:  # 如果没有找到结束标记，使用字符串末尾
                    result = result[start:]
                else:
                    result = result[start:end]
            elif "```" in result:
                # 提取第一个```和```之间的内容
                start = result.find("```") + 3
                end = result.find("```", start)
                if end == -1:  # 如果没有找到结束标记，使用字符串末尾
                    result = result[start:]
                else:
                    result = result[start:end]
            
            result = result.strip()
            parsed = json.loads(result)
            
            if "drawing_info" in parsed:
                return {"drawing_info": parsed["drawing_info"]}
            else:
                return parsed
        except Exception as e:
            logger.error(f"解析LLM响应失败: {e}")
            # 返回空数据结构
            return {"drawing_info": []}  
    def save_data_to_excel(self, data, output_path, dxf_filename):
        """
        将结构化数据保存为Excel文件（支持动态表头，并将数字以数值格式写入）
        """
        logger.info(f"正在保存Excel文件: {output_path}")
        
        wb = Workbook()
        ws = wb.active
        
        # 设置标题
        ws["A1"] = f"{dxf_filename}"
        ws.merge_cells("A1:F3")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(size=14, bold=True)
        
        # 设置所有行的行高为25
        # 先设置标题行（1-3行）
        for row_num in range(1, 4):
            ws.row_dimensions[row_num].height = 25
        
        # 动态生成表头
        drawing_info = data.get("drawing_info", [])
        if drawing_info and isinstance(drawing_info, list) and len(drawing_info) > 0:
            # 从第一个数据项获取所有键作为表头
            first_item = drawing_info[0]
            headers = list(first_item.keys())
            
            # 设置表头（第4行）
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=4, column=col_idx, value=header)
                ws.cell(row=4, column=col_idx).font = Font(bold=True)
            ws.row_dimensions[4].height = 25  # 设置表头行高
            
            # 填充数据（从第5行开始）
            for row_idx, item in enumerate(drawing_info, 5):
                for col_idx, header in enumerate(headers, 1):
                    cell_value = item.get(header, "")
                    # 尝试将值转换为数字
                    final_value = self._convert_to_number_if_possible(cell_value)
                    ws.cell(row=row_idx, column=col_idx, value=final_value)
                ws.row_dimensions[row_idx].height = 25  # 设置数据行行高
        else:
            # 如果没有数据，使用默认表头
            headers = ["图号", "长度", "宽度", "厚度", "数量", "材料", "备注"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=4, column=col_idx, value=header)
                ws.cell(row=4, column=col_idx).font = Font(bold=True)
            ws.row_dimensions[4].height = 25  # 设置表头行高
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 最大宽度限制
        
        # 保存文件
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel文件已保存至: {output_path}")
        return output_path
    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value
    
    def select_dxf_file(self):
        """
        弹出文件选择对话框选择DXF或DWG文件
        """
        root = Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_path = filedialog.askopenfilename(
            title="请选择DXF或DWG文件",
            filetypes=[("CAD文件", "*.dxf *.dwg"), ("DXF文件", "*.dxf"), ("DWG文件", "*.dwg"), ("所有文件", "*.*")]
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return file_path
    
    def select_output_folder(self):
        """
        弹出文件夹选择对话框选择输出文件夹
        """
        root = Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        folder_path = filedialog.askdirectory(
            title="请选择输出文件夹"
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return folder_path
    
    def process_dxf_file(self, dxf_file_path=None, output_folder=None):
        """
        处理DXF或DWG文件的主流程
        """
        try:
            # 如果没有提供文件路径，弹出选择对话框
            if not dxf_file_path:
                dxf_file_path = self.select_dxf_file()
                if not dxf_file_path:
                    logger.info("未选择文件，程序退出")
                    return None
            
            # 检查文件类型，如果是DWG则先转换为DXF
            file_ext = os.path.splitext(dxf_file_path)[1].lower()
            if file_ext == '.dwg':
                logger.info(f"检测到DWG文件，开始转换为DXF: {dxf_file_path}")
                convert_result = self.convert_dwg_to_dxf(dxf_file_path)
                
                if convert_result["status"] != "success":
                    logger.error(f"DWG转换失败: {convert_result['message']}")
                    raise Exception(f"DWG转换失败: {convert_result['message']}")
                
                dxf_file_path = convert_result["dxf_path"]
                logger.info(f"DWG转换成功，使用转换后的DXF文件: {dxf_file_path}")
            elif file_ext != '.dxf':
                raise ValueError(f"不支持的文件格式: {file_ext}。仅支持DXF和DWG文件。")
            
            # 如果没有提供输出文件夹，使用默认位置
            if not output_folder:
                output_folder = os.path.join(current_dir, "output")
            
            # 读取DXF文件中的文本
            dxf_data = self.read_dxf_texts(dxf_file_path)
            
            # 提取文件名
            dxf_filename = os.path.basename(dxf_file_path)
            
            # 使用LLM分析文本内容
            llm_result = self.analyze_text_with_llm(dxf_data['all_text_content'], dxf_filename)
            
            # 生成输出文件名
            output_filename = f"{os.path.splitext(dxf_filename)[0]}_text_analysis.xlsx"
            output_path = os.path.join(output_folder, output_filename)
            
            # 保存到Excel
            excel_path = self.save_data_to_excel(llm_result, output_path, dxf_filename)
            
            logger.info(f"处理完成！Excel文件已保存至: {excel_path}")
            return excel_path
        except Exception as e:
            logger.error(f"处理过程中发生错误: {str(e)}")
            raise

if __name__ == "__main__":
    try:
        converter = DXFTextToExcelConverter()
        print("欢迎使用CAD文件文字信息提取与Excel导出工具")
        print("该工具将读取DXF或DWG文件中的文字信息，使用LLM分析并导出到Excel文件")
        print("支持自动转换DWG文件为DXF格式")
        
        # 处理文件
        excel_file = converter.process_dxf_file()
        
        if excel_file:
            print(f"\n✅ 处理完成！\nExcel文件已保存至: {excel_file}")
    except Exception as e:
        print(f"❌ 处理失败: {str(e)}")
        sys.exit(1)