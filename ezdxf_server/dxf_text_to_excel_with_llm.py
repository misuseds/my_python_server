#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
读取DXF文件的文字信息，使用LLM分析并导出到Excel
支持DWG文件自动转换为DXF后处理
"""
import os
import json
import sys
import re
import subprocess
import time
from ezdxf import readfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 添加父目录到路径，以便导入LLM服务
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# 尝试导入LLM服务类
import traceback
try:
    # 从llm_server导入LLMService
    from llm_server.llm_class import LLMService
    llm_service_available = True
except ImportError as e:
    logger.warning(f"无法导入LLM服务: {e}")
    llm_service_available = False
    logger.debug(traceback.format_exc())

class DXFTextToExcelConverter:
    def __init__(self):
        if llm_service_available:
            self.llm_service = LLMService()
        else:
            self.llm_service = None
            logger.warning("LLM服务不可用，将使用模拟数据进行演示")
        
        # 定义提取字段配置，通过修改这个列表可以同时改变提示语句和JSON示例
        self.extraction_fields = [
            {"key": "drawing_number", "label": "ABC123", "display_name": "图号"},
            {"key": "specification", "label": "PL宽度*长度", "display_name": "规格"},
            
            {"key": "quantity", "label": "5", "display_name": "数量"},
            {"key": "material", "label": "不锈钢", "display_name": "材料"},
          
            {"key": "remark", "label": "不锈钢板", "display_name": "备注"}
        ]
    
    def convert_dwg_to_dxf(self, dwg_file_path):
        """
        使用ODA File Converter将DWG文件转换为DXF文件
        
        Args:
            dwg_file_path (str): DWG文件路径
            
        Returns:
            dict: 转换结果，包含状态和DXF文件路径或错误信息
        """
        try:
            # 检查DWG文件是否存在
            if not os.path.exists(dwg_file_path):
                return {
                    "status": "error",
                    "message": f"DWG文件不存在: {dwg_file_path}"
                }

            # 获取文件目录和文件名
            dwg_dir = os.path.dirname(dwg_file_path)
            dwg_filename = os.path.basename(dwg_file_path)
            dxf_filename = dwg_filename.replace('.dwg', '.dxf')
            dxf_file_path = dwg_file_path.replace('.dwg', '.dxf')
            if dxf_file_path == dwg_file_path:
                dxf_file_path = f"{dwg_file_path}.dxf"

            # 创建输出目录（如果不存在）
            output_dir = os.path.dirname(dxf_file_path)
            os.makedirs(output_dir, exist_ok=True)

            # 构建命令
            cmd = [
                "ODAFileConverter",
                dwg_dir,
                output_dir,
                "ACAD2018",   # 输入版本
                "DXF",        # 输出格式
                "0",          # 非递归
                "1",          # 执行审计
                "*.DWG"       # 只处理DWG文件
            ]

            logger.info(f"执行DWG转换命令: {' '.join(cmd)}")

            # 执行转换命令
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=120)

            if result.returncode == 0:
                # 检查DXF文件是否生成
                if os.path.exists(dxf_file_path):
                    response = {
                        "status": "success",
                        "message": f"DWG文件已成功转换为DXF: {dxf_file_path}",
                        "dxf_path": dxf_file_path
                    }
                    return response
                else:
                    response = {
                        "status": "error",
                        "message": f"转换完成但未生成DXF文件。请检查输出目录: {output_dir}"
                    }
                    return response
            else:
                response = {
                    "status": "error",
                    "message": f"ODA转换失败: {result.stderr}"
                }
                return response

        except subprocess.TimeoutExpired:
            return {"status": "error", "message": "转换超时"}
        except FileNotFoundError:
            return {"status": "error", "message": "未找到ODAFileConverter命令，请确保ODA File Converter已正确安装并加入PATH"}
        except Exception as e:
            return {"status": "error", "message": f"转换过程中发生错误: {str(e)}"}
    
    def read_dxf_texts(self, dxf_file_path):
        """
        读取DXF文件中的所有文本实体，包括标注文字
        """
        logger.info(f"开始读取DXF文件: {dxf_file_path}")
        
        # 检查文件是否存在
        if not os.path.exists(dxf_file_path):
            raise FileNotFoundError(f"文件不存在: {dxf_file_path}")
        
        # 检查文件格式
        file_ext = os.path.splitext(dxf_file_path)[1].lower()
        if file_ext != '.dxf':
            raise ValueError(f"不支持的文件格式: {file_ext}。仅支持DXF文件。")
        
        try:
            # 读取DXF文件
            doc = readfile(dxf_file_path)
            msp = doc.modelspace()
            
            texts = []
            for entity in msp:
                if entity.dxftype() == 'TEXT':
                    texts.append({
                        'content': entity.dxf.text,
                        'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                    })
                elif entity.dxftype() == 'MTEXT':
                    texts.append({
                        'content': entity.text,
                        'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                    })
                elif entity.dxftype() == 'DIMENSION':
                    # 处理标注实体
                    try:
                        # 获取标注文字
                        dim_text = entity.dxf.text
                        if dim_text and dim_text.strip():
                            texts.append({
                                'content': dim_text,
                                'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                            })
                        # 如果标注文字为空，尝试获取测量值
                        elif entity.dim_text:
                            texts.append({
                                'content': entity.dim_text,
                                'position': {'x': entity.dxf.insert.x, 'y': entity.dxf.insert.y}
                            })
                    except Exception as e:
                        logger.debug(f"处理标注实体时出错: {e}")
                        continue
            
            all_text_content = '\n'.join([text['content'] for text in texts])
            logger.info(f"成功读取DXF文件，共提取 {len(texts)} 个文本实体")
            
            return {
                'status': 'success',
                'texts': texts,
                'all_text_content': all_text_content,
                'text_count': len(texts),
                'file_path': dxf_file_path
            }
        except Exception as e:
            logger.error(f"处理文件时出错: {str(e)}")
            raise
    
    def analyze_text_with_llm(self, text_content, file_name):
        """
        使用LLM分析文本内容
        """
        # 根据当前字段配置动态生成提示语句
        labels = [field["display_name"] for field in self.extraction_fields]
        labels_str = "、".join(labels)
        instruction = f"请分析以下DXF文件中的文本信息，提取出{labels_str}等关键信息,不要统计，并以JSON格式返回。"
        
        # 动态构建JSON示例中的字段部分
        field_lines = []
        for field in self.extraction_fields:
            field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
        
        json_fields_block = ",\n".join(field_lines)
        
        logger.info("正在使用LLM分析文本内容")
        
        messages = [
            {
                "role": "user",
                "content": f'''{instruction}
文本内容：
{text_content}

输出格式要求：
{{
    "drawing_info": [
        {{
{json_fields_block}
        }}
    ]
}}'''
            }
        ]
        
        try:
            # 调用LLM服务进行文本分析
            result = self.llm_service.create(messages)
            logger.info(f"LLM分析完成")
            print(f"LLM原始返回: {result}")
            
            # 解析LLM返回结果
            return self.parse_llm_response(result)
        except Exception as e:
            logger.error(f"LLM分析出错: {str(e)}")
            raise
    
    def parse_llm_response(self, response_data):
        """
        解析LLM返回的结构化数据
        """
        try:
            result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            # 使用正则表达式提取JSON部分
            # 匹配 ```json 或 ``` 代码块中的内容
            import re
            json_match = re.search(r'```(?:json)?\s*({.*?})\s*```', result, re.DOTALL)
            if json_match:
                result = json_match.group(1)
            else:
                # 如果没有找到代码块，尝试找到第一个 { 和最后一个 } 之间的内容
                first_brace = result.find('{')
                last_brace = result.rfind('}')
                if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
                    result = result[first_brace:last_brace+1]
            
            result = result.strip()
            parsed = json.loads(result)
            
            if "drawing_info" in parsed:
                return {"drawing_info": parsed["drawing_info"]}
            else:
                return {}
        except Exception as e:
            print(f"解析失败: {e}")
            # 调试信息，帮助排查问题
            print(f"LLM返回的原始内容: {result}")
            return {}

    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 处理Unicode数字字符（如①②③等）
            try:
                unicode_digits = {
                    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
                    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10
                }
                if value in unicode_digits:
                    return unicode_digits[value]
            except:
                pass
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value

    def save_data_to_excel(self, data, output_path="output.xlsx", dxf_filename=None):
        """
        将结构化数据保存为Excel文件
        
        Args:
            data (dict): 结构化数据
            output_path (str): 输出Excel文件路径
            dxf_filename (str): 原始DXF文件名，用于填入Excel标题
        """
        # 防止文件名重复
        base_path, ext = os.path.splitext(output_path)
        counter = 1
        final_output_path = output_path
        while os.path.exists(final_output_path):
            final_output_path = f"{base_path}_{counter}{ext}"
            counter += 1
        
        wb = Workbook()
        ws = wb.active
        
        # 设置标题
        if dxf_filename:
            ws["A1"] = dxf_filename
            ws.merge_cells("A1:E1")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(size=14, bold=True)
            ws.row_dimensions[1].height = 25
            start_row = 2
        else:
            start_row = 1
        
        rows = []
        
        if "drawing_info" in data and data["drawing_info"]:
            # 创建显示名称映射
            display_names = {field["key"]: field["display_name"] for field in self.extraction_fields}
            
            # 创建表头行（中文）
            header_row = []
            field_order = []  # 保持列顺序
            
            # 按预定义顺序添加字段
            for field_config in self.extraction_fields:
                field_key = field_config["key"]
                header_row.append(display_names.get(field_key, field_key))
                field_order.append(field_key)
                
            rows.append(header_row)
            
            # 填充数据行
            for item in data["drawing_info"]:
                data_row = []
                for field in field_order:
                    cell_value = item.get(field, "")
                    # 尝试将值转换为数字
                    final_value = self._convert_to_number_if_possible(cell_value)
                    data_row.append(final_value)
                rows.append(data_row)
        
        # 写入Excel
        for i, row in enumerate(rows, start=start_row):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
            # 设置每行行高为25
            ws.row_dimensions[i].height = 25
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大宽度
        
        wb.save(final_output_path)
        print(f"Excel 文件已保存至: {final_output_path}")
        return final_output_path

    def select_dxf_files(self):
        """
        弹出文件选择窗口选择多个DXF或DWG文件
        """
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        root.attributes('-topmost', True)  # 确保对话框在最前面
        
        file_types = [
            ("All supported files", "*.dxf *.dwg"),
            ("DXF files", "*.dxf"),
            ("DWG files", "*.dwg"),
            ("All files", "*.*")
        ]
        
        file_paths = filedialog.askopenfilenames(
            title="请选择一个或多个DXF或DWG文件",
            filetypes=file_types
        )
        
        root.destroy()  # 销毁tkinter根窗口
        return list(file_paths)

    def process_dxf_file(self, dxf_file_path):
        """
        处理单个DXF或DWG文件
        
        Args:
            dxf_file_path (str): 输入文件路径（DXF或DWG）
        """
        try:
            # 检查文件类型，如果是DWG则先转换为DXF
            file_ext = os.path.splitext(dxf_file_path)[1].lower()
            original_file_path = dxf_file_path  # 保存原始文件路径用于确定输出目录
            
            if file_ext == '.dwg':
                logger.info(f"检测到DWG文件，开始转换为DXF: {dxf_file_path}")
                convert_result = self.convert_dwg_to_dxf(dxf_file_path)
                
                if convert_result["status"] != "success":
                    logger.error(f"DWG转换失败: {convert_result['message']}")
                    raise Exception(f"DWG转换失败: {convert_result['message']}")
                
                dxf_file_path = convert_result["dxf_path"]
                logger.info(f"DWG转换成功，使用转换后的DXF文件: {dxf_file_path}")
            elif file_ext != '.dxf':
                raise ValueError(f"不支持的文件格式: {file_ext}。仅支持DXF和DWG文件。")
            
            # 获取文件目录和文件名（不含扩展名）
            file_dir = os.path.dirname(original_file_path)
            file_name = os.path.splitext(os.path.basename(original_file_path))[0]
            output_path = os.path.join(file_dir, f"{file_name}.xlsx")
            
            # 读取DXF文件中的文本
            dxf_data = self.read_dxf_texts(dxf_file_path)
            
            # 使用LLM分析文本内容
            llm_result = self.analyze_text_with_llm(dxf_data['all_text_content'], os.path.basename(dxf_file_path))
            
            # 保存到Excel
            self.save_data_to_excel(llm_result, output_path, os.path.basename(original_file_path))
        except Exception as e:
            print("错误:", str(e))

    def batch_process_files(self, file_paths):
        """
        批量处理多个文件，每个文件生成单独的Excel文件
        
        Args:
            file_paths (list): 文件路径列表
        """
        if not file_paths:
            print("未选择任何文件")
            return
        
        for i, file_path in enumerate(file_paths, 1):
            print(f"\n正在处理第 {i}/{len(file_paths)} 个文件: {file_path}")
            
            if not os.path.exists(file_path):
                print(f"警告: 文件 '{file_path}' 不存在，跳过处理")
                continue
                
            try:
                self.process_dxf_file(file_path)
                print(f"完成处理: {file_path}")
            except Exception as e:
                print(f"处理文件 '{file_path}' 时出现错误: {e}")

# 创建全局转换器实例
converter = DXFTextToExcelConverter()

def main():
    # 直接选择多个文件进行批量处理
    file_paths = converter.select_dxf_files()
    
    if not file_paths:
        print("未选择文件，程序退出")
        sys.exit(0)
        
    try:
        converter.batch_process_files(file_paths)
        print("\n所有文件处理完成！")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")

if __name__ == "__main__":
    try:
        print("欢迎使用CAD文件文字信息提取与Excel导出工具")
        print("该工具将读取DXF或DWG文件中的文字信息，使用LLM分析并导出到Excel文件")
        print("支持自动转换DWG文件为DXF格式")
        main()
    except Exception as e:
        print(f"❌ 处理失败: {str(e)}")
        sys.exit(1)